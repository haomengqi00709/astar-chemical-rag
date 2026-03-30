"""
PM Agent — Project Manager (v2)

Reads a client SOW and produces a project_context.json using:
  1. catalog.json  — routing layer  (what work needs to happen)
  2. RAG           — knowledge layer (what the standards require)
  3. Gemini        — reasoning layer (extract, match, explain)

Also generates filled deliverable documents for all PM-owned items in the
document register (Document Register, PEP, Project Schedule), saved to
deliverables/ and embedded in the returned context as context['deliverables'].

Thinking is shown step by step before anything is written.
Human confirmation is required before project_context.json is saved.

Usage:
    python work_agents/pm_agent.py --demo
    python work_agents/pm_agent.py --sow path/to/sow.txt --end-date 2026-06-01
    python work_agents/pm_agent.py --demo --save work_agents/project_context.json
    python work_agents/pm_agent.py --demo --json   # raw JSON, no interaction
"""

import argparse
import json
import os
import re
import sys
import zipfile
from datetime import datetime, timedelta
from pathlib import Path

import chromadb
from dotenv import load_dotenv
from google import genai
from google.genai import types

load_dotenv()

AGENT_DIR       = Path(__file__).parent                  # Daniel - Project Manager /
WORK_AGENTS     = AGENT_DIR.parent                        # work_agents/
RAG_ROOT        = WORK_AGENTS.parent                      # project root
CHROMA_PATH     = RAG_ROOT / 'chroma_db'
CATALOG_PATH    = WORK_AGENTS / 'catalog.json'
COLLECTION_NAME = 'compliance_docs'
EMBEDDING_MODEL = 'models/gemini-embedding-001'
GENERATION_MODEL = 'gemini-2.5-flash'

SAMPLE_SOW = """
Project Title: MEG Transfer Pump Skid
Project Number: BSC-2026-001
Client: B Star Chemetics
Location: Sarnia, Ontario, Canada

Scope of Work:
Design, procure, and commission a MEG transfer pump skid consisting of one (1) duty and
one (1) standby horizontal centrifugal pump to transfer monoethylene glycol (MEG) from
an intermediate flash separator to the MEG regeneration column at elevated pressure.
The scope includes hydraulic design and pump sizing per 4-PRC-0007, pump and motor
selection in accordance with API 610, preparation of pump calculation sheet (4-CAL-0001),
pump data sheet (4-DST-0001), process design basis and fluid property assessment per
1-PRC-0001, document register and project execution plan.

Key design parameters:
- Fluid: Monoethylene Glycol (MEG), 95% by weight
- Normal flow rate: 45 m³/h
- Design pressure: 10 bar(g)
- Operating temperature: 65°C
- Fluid density: 1095 kg/m³
- Suction vessel pressure: 1.5 barg
- Discharge vessel pressure: 3.2 barg
- Suction static head: 3.0 m
- Discharge static head: 8.0 m
- Suction-side friction losses: 0.8 m
- Discharge-side friction losses: 2.5 m

Equipment: 1 × horizontal centrifugal pump, API 610 OH2 (duty) + 1 × standby, electric
motor driven, TEFC enclosure, mechanical seal API Plan 11.
"""


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def emit_progress(step: int, title: str, status: str = 'running', data: dict | None = None):
    """Write a progress event to stderr as JSON (consumed by server.js for SSE)."""
    msg = {'step': step, 'title': title, 'status': status}
    if data:
        msg['data'] = data
    sys.stderr.write(json.dumps(msg) + '\n')
    sys.stderr.flush()


def read_sow_file(path: str) -> str:
    p = Path(path)
    if p.suffix.lower() == '.docx':
        z = zipfile.ZipFile(p)
        xml = z.read('word/document.xml').decode('utf-8', errors='ignore')
        text = re.sub(r'<[^>]+>', '', xml)
        return re.sub(r'\s+', ' ', text).strip()
    return p.read_text(encoding='utf-8', errors='ignore')


def load_catalog() -> dict:
    if not CATALOG_PATH.exists():
        raise FileNotFoundError(
            f'catalog.json not found at {CATALOG_PATH}. '
            'Run: python work_agents/build_catalog.py'
        )
    return json.loads(CATALOG_PATH.read_text())


def query_rag(question: str, collection, client: genai.Client, top_k: int = 5) -> str:
    resp = client.models.embed_content(
        model=EMBEDDING_MODEL,
        contents=question,
        config=types.EmbedContentConfig(task_type='RETRIEVAL_QUERY'),
    )
    results = collection.query(
        query_embeddings=[resp.embeddings[0].values],
        n_results=top_k,
        where={'authority': 'Standard'},
    )
    chunks = []
    for i, text in enumerate(results['documents'][0]):
        meta = results['metadatas'][0][i]
        chunks.append(f"[{meta.get('doc_id', '?')}] {text[:200]}")
    return '\n'.join(chunks)


def calculate_date(end_date: datetime, weeks_before: float) -> str:
    return (end_date - timedelta(weeks=weeks_before)).strftime('%Y-%m-%d')


# ─────────────────────────────────────────────────────────────────────────────
# Step 1 — Extract project metadata from SOW
# ─────────────────────────────────────────────────────────────────────────────

def extract_project_metadata(sow_text: str, client: genai.Client) -> dict:
    prompt = f"""Extract structured project metadata from this engineering Scope of Work.
Return ONLY a valid JSON object with these exact keys (use null if not found):
{{
  "client": "company name",
  "project_title": "short project title",
  "location": "city, country",
  "equipment_type": "primary equipment type (e.g. centrifugal pump)",
  "fluid": "fluid name and concentration",
  "flow_rate_m3h": <number or null>,
  "design_pressure_bar": <number or null>,
  "temperature_c": <number or null>,
  "fluid_density_kgm3": <number or null>,
  "suction_pressure_barg": <number or null>,
  "discharge_pressure_barg": <number or null>,
  "suction_static_head_m": <number or null>,
  "discharge_static_head_m": <number or null>,
  "suction_friction_loss_m": <number or null>,
  "discharge_friction_loss_m": <number or null>,
  "special_requirements": "any notable constraints or standards mentioned"
}}

SOW:
{sow_text}"""

    resp = client.models.generate_content(
        model=GENERATION_MODEL,
        contents=prompt,
        config=types.GenerateContentConfig(temperature=0.1),
    )
    text = resp.text.strip()
    if text.startswith('```'):
        text = text.split('```')[1]
        if text.startswith('json'):
            text = text[4:]
    return json.loads(text.strip())


# ─────────────────────────────────────────────────────────────────────────────
# Step 2 — Match project type from catalog
# ─────────────────────────────────────────────────────────────────────────────

def match_project_type(summary: dict, catalog: dict) -> tuple[str, dict] | tuple[None, None]:
    equipment = (summary.get('equipment_type') or '').lower()
    fluid     = (summary.get('fluid') or '').lower()
    combined  = f"{equipment} {fluid}"

    best_match = None
    best_score = 0

    for ptype, pdata in catalog['project_types'].items():
        score = sum(
            1 for trigger in pdata.get('triggers', [])
            if trigger.lower() in combined
        )
        if score > best_score:
            best_score = score
            best_match = ptype

    if best_match:
        return best_match, catalog['project_types'][best_match]
    return None, None


# ─────────────────────────────────────────────────────────────────────────────
# Step 3 — Evaluate risk flags
# ─────────────────────────────────────────────────────────────────────────────

def evaluate_risk_flags(summary: dict, ptype_data: dict,
                        collection, client: genai.Client) -> list[dict]:
    """Check each risk flag. If fired, run RAG query and attach result."""
    fluid    = (summary.get('fluid') or '').lower()
    pressure = summary.get('design_pressure_bar') or 0
    temp     = summary.get('temperature_c') or 0

    fired = []
    for flag in ptype_data.get('risk_flags', []):
        triggered = False
        trigger_reason = ''

        # Keyword match against fluid / special requirements
        sow_text = f"{fluid} {(summary.get('special_requirements') or '').lower()}"
        for kw in flag.get('sow_keywords', []):
            if kw.lower() in sow_text:
                triggered = True
                trigger_reason = f'keyword "{kw}" found in SOW'
                break

        # Numeric threshold checks
        thresholds = flag.get('sow_thresholds', {})
        if not triggered and thresholds.get('pressure_above') and pressure > thresholds['pressure_above']:
            triggered = True
            trigger_reason = f'design pressure {pressure} bar > {thresholds["pressure_above"]} bar threshold'
        if not triggered and thresholds.get('temp_below') and temp < thresholds['temp_below']:
            triggered = True
            trigger_reason = f'temperature {temp}°C below {thresholds["temp_below"]}°C threshold'

        if triggered:
            rag_result = ''
            if flag.get('rag_query'):
                rag_result = query_rag(flag['rag_query'], collection, client, top_k=3)
            fired.append({
                **flag,
                'trigger_reason': trigger_reason,
                'rag_result': rag_result,
            })

    return fired


# ─────────────────────────────────────────────────────────────────────────────
# Step 4 — Build document register with dates
# ─────────────────────────────────────────────────────────────────────────────

def build_document_register(ptype_data: dict, fired_flags: list[dict],
                             end_date: datetime) -> list[dict]:
    # Start from catalog required_docs
    docs = [dict(d) for d in ptype_data['required_docs']]
    existing_ids = {d['doc_id'] for d in docs}

    # Add extra docs from fired risk flags
    for flag in fired_flags:
        for extra_id in flag.get('additional_docs', []):
            if extra_id not in existing_ids:
                docs.append({
                    'doc_id': extra_id,
                    'type': 'SPC',
                    'title': f'Additional document — {flag["condition"]}',
                    'discipline': 4,
                    'discipline_name': 'Equipment',
                    'assigned_to': 'Mechanical Engineer',
                    'reason': f'Added due to risk flag: {flag["condition"]}',
                    'depends_on': [],
                    'weeks_before_end': 2,
                })
                existing_ids.add(extra_id)

    # Calculate planned dates
    for doc in docs:
        weeks = doc.get('weeks_before_end', 2)
        doc['planned_date'] = calculate_date(end_date, weeks)
        doc['status'] = 'Not Started'

    return docs


# ─────────────────────────────────────────────────────────────────────────────
# Step 5 — Build task assignments
# ─────────────────────────────────────────────────────────────────────────────

def build_task_assignments(summary: dict, register: list[dict]) -> dict:
    process_docs    = [d for d in register if d['assigned_to'] == 'Process Engineer']
    mechanical_docs = [d for d in register if d['assigned_to'] == 'Mechanical Engineer']
    process_ids     = [d['doc_id'] for d in process_docs]
    equipment       = summary.get('equipment_type', 'equipment')
    fluid           = summary.get('fluid', 'TBD')
    flow            = summary.get('flow_rate_m3h', 'TBD')

    return {
        'process_task': {
            'agent': 'Process Engineer',
            'deliverables': process_docs,
            'project_summary': summary,
            'instructions': (
                f"Complete process engineering deliverables for the {equipment} project. "
                f"Fluid: {fluid}. Flow rate: {flow} m³/h. "
                "1. Identify fluid code from 1-LST-0002. "
                "2. Fill in 1-PRC-0001 Process Design Criteria. "
                "3. Produce 1-CAL-6510 process calculation summary."
            ),
        },
        'mechanical_task': {
            'agent': 'Mechanical Engineer',
            'deliverables': mechanical_docs,
            'project_summary': summary,
            'wait_for': process_ids,
            'instructions': (
                f"Perform pump engineering deliverables for the {equipment} project. "
                "Wait for Process Calculations before starting. "
                "1. Run full pump hydraulic calculation using Pump_Calculation_Template. "
                "2. Check corrosion allowance against 5-LST-0003. "
                "3. Generate pump datasheet (4-DST-XXXX). "
                "4. Assemble technical bid package."
            ),
        },
    }


# ─────────────────────────────────────────────────────────────────────────────
# Display — thinking chain
# ─────────────────────────────────────────────────────────────────────────────

def print_thinking(summary: dict, project_type: str, ptype_data: dict,
                   fired_flags: list[dict], register: list[dict], end_date_str: str):
    sep = '─' * 64

    print(f'\n{"="*64}')
    print('  PM AGENT — THINKING')
    print(f'{"="*64}')

    # SOW analysis
    print(f'\n  [{1}] SOW ANALYSIS')
    print(f'  {sep}')
    print(f'  Client   : {summary.get("client", "N/A")}')
    print(f'  Project  : {summary.get("project_title", "N/A")}')
    print(f'  Location : {summary.get("location", "N/A")}')
    print(f'  Equipment: {summary.get("equipment_type", "N/A")}')
    print(f'  Fluid    : {summary.get("fluid", "N/A")}')
    print(f'  Flow     : {summary.get("flow_rate_m3h")} m³/h  |  '
          f'P: {summary.get("design_pressure_bar")} bar  |  '
          f'T: {summary.get("temperature_c")} °C')
    if summary.get('special_requirements'):
        print(f'  Special  : {summary["special_requirements"]}')

    # Catalog match
    print(f'\n  [{2}] CATALOG MATCH')
    print(f'  {sep}')
    triggers_hit = [t for t in ptype_data.get('triggers', [])
                    if t.lower() in f"{summary.get('equipment_type','')} {summary.get('fluid','')}".lower()]
    print(f'  Matched project type : {project_type}')
    print(f'  Triggers matched     : {", ".join(triggers_hit)}')
    print(f'  Agents involved      : {", ".join(ptype_data.get("agents", []))}')
    print(f'  Base documents loaded: {len(ptype_data["required_docs"])}')

    # Risk flags
    print(f'\n  [{3}] RISK ANALYSIS')
    print(f'  {sep}')
    if not fired_flags:
        print('  No risk flags triggered.')
    for flag in fired_flags:
        print(f'  ⚠ FLAG: {flag["condition"]}')
        print(f'    Triggered by : {flag["trigger_reason"]}')
        print(f'    Warning      : {flag["warning"]}')
        if flag.get('additional_docs'):
            print(f'    Adding docs  : {", ".join(flag["additional_docs"])}')
        if flag.get('rag_result'):
            print(f'    RAG finding  :')
            for line in flag['rag_result'].split('\n')[:3]:
                if line.strip():
                    print(f'      {line.strip()[:100]}')

    # Proposed document register
    print(f'\n  [{4}] PROPOSED DOCUMENT REGISTER  ({len(register)} documents, end date {end_date_str})')
    print(f'  {sep}')
    print(f'  {"Doc ID":<18} {"Owner":<22} {"Due Date":<12} {"Depends On":<20} Title')
    print(f'  {"─"*18} {"─"*22} {"─"*12} {"─"*20} {"─"*28}')
    for doc in register:
        deps = ', '.join(doc.get('depends_on', [])) or '—'
        print(f'  {doc["doc_id"]:<18} {doc["assigned_to"]:<22} '
              f'{doc.get("planned_date",""):<12} {deps:<20} {doc["title"]}')
        print(f'  {"":18}   → {doc["reason"][:80]}')

    print()


# ─────────────────────────────────────────────────────────────────────────────
# Step 8 — Generate PM-owned deliverable documents
# ─────────────────────────────────────────────────────────────────────────────

DELIVERABLES_DIR = AGENT_DIR / 'deliverables'

# Historical template files for specific PM documents
PM_TEMPLATE_FILES: dict[str, Path] = {
    '0-LST-0001': AGENT_DIR / '0-LST-0001-1-R0 [Template] Document Register.xlsx',
}

# RAG queries per document type — what to ask the knowledge base
PM_DOC_RAG_QUERIES: dict[str, str] = {
    'LST': 'What are the required fields and format for a project document register list?',
    'PRC': 'What sections should a Project Execution Plan contain for an engineering project?',
    'DAT': 'What key milestones and format should a project schedule contain?',
}


def read_template_file(path: Path) -> str:
    """Read a template/reference file and return its content as plain text."""
    try:
        suffix = path.suffix.lower()
        if suffix in ('.xlsx', '.xlsm'):
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            lines = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lines.append(f'[Sheet: {sheet}]')
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        lines.append('  ' + ' | '.join(cells))
            return '\n'.join(lines)[:2500]
        elif suffix == '.docx':
            import zipfile as _zf
            z = _zf.ZipFile(path)
            xml = z.read('word/document.xml').decode('utf-8', errors='ignore')
            text = re.sub(r'<[^>]+>', '', xml)
            return re.sub(r'\s+', ' ', text).strip()[:2500]
    except Exception as e:
        return f'[Template read error: {e}]'
    return ''


def generate_deliverable_content(doc: dict, context: dict,
                                  rag_snippets: str, template_text: str,
                                  client: genai.Client) -> str:
    """Use Gemini to generate a filled document from project data + RAG + template."""
    summary  = context['project_summary']
    register = context['document_register']
    end_date = context['project_end_date']
    flags    = context.get('risk_flags_fired', [])

    project_str = (
        f"Project: {summary.get('project_title')}\n"
        f"Client: {summary.get('client')}  |  Location: {summary.get('location')}\n"
        f"Equipment: {summary.get('equipment_type')}\n"
        f"Fluid: {summary.get('fluid')}  |  Flow: {summary.get('flow_rate_m3h')} m³/h  |  "
        f"Pressure: {summary.get('design_pressure_bar')} bar  |  Temp: {summary.get('temperature_c')} °C\n"
        f"Project End Date: {end_date}\n"
        f"Risk Flags Active: {', '.join(f['condition'] for f in flags) if flags else 'None'}"
    )

    register_str = '\n'.join(
        f"  {d['doc_id']:<18} {d['assigned_to']:<22} {d.get('planned_date', ''):<12} {d['title']}"
        for d in register
    )

    rag_section      = f"\n--- RELEVANT STANDARDS (from RAG knowledge base) ---\n{rag_snippets}\n" if rag_snippets else ""
    template_section = f"\n--- HISTORICAL TEMPLATE / REFERENCE FILE ---\n{template_text}\n" if template_text else ""

    prompt = f"""You are a senior Project Manager generating a formal engineering project deliverable.

DOCUMENT TO GENERATE:
  ID:       {doc['doc_id']}
  Title:    {doc['title']}
  Type:     {doc.get('type')} ({doc.get('discipline_name', '')})
  Due Date: {doc.get('planned_date', 'TBD')}
  Purpose:  {doc.get('reason', '')}

PROJECT DATA:
{project_str}

FULL DOCUMENT REGISTER ({len(register)} deliverables on this project):
{register_str}
{rag_section}{template_section}
Generate a complete, professional version of this document filled with the actual project data above.
Use the RAG standards and template reference to ensure correct structure and content.
Format in markdown with clear headers and tables where appropriate.
The first line must be exactly:  # {doc['doc_id']} — {doc['title']}"""

    resp = client.models.generate_content(
        model=GENERATION_MODEL,
        contents=prompt,
        config=types.GenerateContentConfig(temperature=0.2),
    )
    return resp.text.strip()


def generate_pm_deliverables(register: list[dict], context: dict,
                              collection, client: genai.Client,
                              json_mode: bool = False) -> dict[str, str]:
    """Generate filled documents for all PM-owned deliverables.
    Returns dict of {doc_id: markdown_content}.
    """
    DELIVERABLES_DIR.mkdir(exist_ok=True)
    pm_docs   = [d for d in register if d.get('assigned_to') == 'PM']
    generated: dict[str, str] = {}

    for i, doc in enumerate(pm_docs):
        doc_id   = doc['doc_id']
        doc_type = doc.get('type', 'GEN')

        if json_mode:
            emit_progress(7, f'Generating {doc_id}',
                          data={'doc_id': doc_id, 'index': i + 1, 'total': len(pm_docs)})
        else:
            print(f'  Generating {doc_id} — {doc["title"]}...')

        try:
            # 1. Query RAG for standards relevant to this document type
            rag_query    = PM_DOC_RAG_QUERIES.get(doc_type, f'Requirements for {doc["title"]}')
            rag_snippets = query_rag(rag_query, collection, client, top_k=4)

            # 2. Load historical template/reference if one exists for this doc_id
            template_text = ''
            tpl_path = PM_TEMPLATE_FILES.get(doc_id)
            if tpl_path and tpl_path.exists():
                template_text = read_template_file(tpl_path)

            # 3. Generate content with Gemini
            content = generate_deliverable_content(doc, context, rag_snippets, template_text, client)

            # 4. Save .md file locally for reference
            safe_id  = doc_id.replace('/', '_').replace('.', '_')
            filepath = DELIVERABLES_DIR / f'{safe_id}.md'
            filepath.write_text(content, encoding='utf-8')

            generated[doc_id] = content

        except Exception as e:
            generated[doc_id] = f'[Generation failed: {e}]'

        if not json_mode:
            ok = '✓' if not generated[doc_id].startswith('[Generation failed') else '✗'
            print(f'  {ok} {doc_id}')

    return generated


# ─────────────────────────────────────────────────────────────────────────────
# Handoff brief — Gemini writes a plain-language summary of the plan
# ─────────────────────────────────────────────────────────────────────────────

def generate_handoff_brief(context: dict, client: genai.Client) -> str:
    summary   = context['project_summary']
    register  = context['document_register']
    tasks     = context['task_assignments']
    flags     = context['risk_flags_fired']
    end_date  = context['project_end_date']

    process_docs    = [f"{d['doc_id']} — {d['title']} (due {d['planned_date']})"
                       for d in tasks['process_task']['deliverables']]
    mechanical_docs = [f"{d['doc_id']} — {d['title']} (due {d['planned_date']})"
                       for d in tasks['mechanical_task']['deliverables']]
    wait_for        = tasks['mechanical_task'].get('wait_for', [])
    flag_lines      = [f"- {f['condition']}: {f['warning']}" for f in flags]

    prompt = f"""You are a senior Project Manager writing a concise internal handoff brief for an engineering project.

Project details:
- Client: {summary.get('client')}
- Project: {summary.get('project_title')}
- Location: {summary.get('location')}
- Equipment: {summary.get('equipment_type')}
- Fluid: {summary.get('fluid')}
- Flow: {summary.get('flow_rate_m3h')} m³/h  |  Pressure: {summary.get('design_pressure_bar')} bar  |  Temp: {summary.get('temperature_c')} °C
- Project end date: {end_date}
- Special requirements: {summary.get('special_requirements')}

Process Engineer deliverables:
{chr(10).join(process_docs)}

Mechanical Engineer deliverables:
{chr(10).join(mechanical_docs)}
(Mechanical must wait for: {', '.join(wait_for)} before starting pump calculations)

Active risk flags:
{chr(10).join(flag_lines) if flag_lines else '- None'}

Write a clear, professional handoff brief with these sections:
1. Project Overview — 2-3 sentences summarising what this project is and what needs to be delivered
2. Process Engineer Tasks — what they need to do, what inputs they work from, and what they hand off to Mechanical
3. Mechanical Engineer Tasks — what they need to do, what they are waiting for from Process, and what the final deliverables are
4. Risk Flags & Engineering Implications — for each active flag, explain what it means for the work (material selection, testing requirements, additional documents needed)
5. Key Dates — a short ordered list of the critical milestones

Write in plain English. Be specific. Do not use bullet points for everything — use short paragraphs where appropriate."""

    resp = client.models.generate_content(
        model=GENERATION_MODEL,
        contents=prompt,
        config=types.GenerateContentConfig(temperature=0.3),
    )
    return resp.text.strip()


# ─────────────────────────────────────────────────────────────────────────────
# Human confirmation
# ─────────────────────────────────────────────────────────────────────────────

def ask_confirmation(json_mode: bool) -> bool:
    if json_mode:
        return True  # auto-confirm in JSON mode
    print('─' * 64)
    print('  Do you confirm this plan? (yes / no)')
    print('  Type "yes" to save project_context.json and hand off to agents.')
    print('─' * 64)
    answer = input('  > ').strip().lower()
    return answer in ('yes', 'y', 'confirm')


# ─────────────────────────────────────────────────────────────────────────────
# Main pipeline
# ─────────────────────────────────────────────────────────────────────────────

def run_pm_agent(sow_text: str, end_date_str: str,
                 json_mode: bool = False) -> dict | None:

    api_key = os.environ.get('GOOGLE_API_KEY')
    if not api_key:
        raise RuntimeError('Set GOOGLE_API_KEY environment variable.')

    client     = genai.Client(api_key=api_key)
    chroma     = chromadb.PersistentClient(path=str(CHROMA_PATH))
    collection = chroma.get_collection(COLLECTION_NAME)
    end_date   = datetime.strptime(end_date_str, '%Y-%m-%d')

    if not json_mode:
        print('\nLoading catalog...')
    if json_mode:
        emit_progress(0, 'Loading catalog')
    catalog = load_catalog()

    # ── Step 1: Extract metadata ──────────────────────────────────────────────
    if not json_mode:
        print('Analysing SOW...')
    if json_mode:
        emit_progress(1, 'Analysing SOW')
    summary = extract_project_metadata(sow_text, client)
    if json_mode:
        emit_progress(1, 'SOW Analysis', status='done', data={
            'client':    summary.get('client'),
            'project':   summary.get('project_title'),
            'location':  summary.get('location'),
            'equipment': summary.get('equipment_type'),
            'fluid':     summary.get('fluid'),
            'flow':      summary.get('flow_rate_m3h'),
            'pressure':  summary.get('design_pressure_bar'),
            'temp':      summary.get('temperature_c'),
        })

    # ── Step 2: Match project type ────────────────────────────────────────────
    if not json_mode:
        print('Matching project type from catalog...')
    if json_mode:
        emit_progress(2, 'Matching project type')
    project_type, ptype_data = match_project_type(summary, catalog)

    if not project_type:
        print('ERROR: Could not match SOW to any project type in catalog.')
        print(f'Available types: {list(catalog["project_types"].keys())}')
        print(f'Equipment extracted: {summary.get("equipment_type")}')
        return None

    if json_mode:
        triggers_hit = [t for t in ptype_data.get('triggers', [])
                        if t.lower() in f"{summary.get('equipment_type','')} {summary.get('fluid','')}".lower()]
        emit_progress(2, 'Catalog Match', status='done', data={
            'project_type':   project_type,
            'triggers_hit':   triggers_hit,
            'agents':         ptype_data.get('agents', []),
            'base_doc_count': len(ptype_data.get('required_docs', [])),
        })

    # ── Step 3: Evaluate risk flags (with RAG) ────────────────────────────────
    if not json_mode:
        print('Evaluating risk flags...')
    if json_mode:
        emit_progress(3, 'Evaluating risk flags (querying RAG)')
    fired_flags = evaluate_risk_flags(summary, ptype_data, collection, client)
    if json_mode:
        emit_progress(3, 'Risk Analysis', status='done', data={
            'flags_fired': [
                {'condition': f['condition'], 'warning': f['warning'],
                 'trigger_reason': f['trigger_reason']}
                for f in fired_flags
            ] if fired_flags else [],
            'total_checked': len(ptype_data.get('risk_flags', [])),
        })

    # ── Step 4: Build document register ──────────────────────────────────────
    if json_mode:
        emit_progress(4, 'Building document register')
    register = build_document_register(ptype_data, fired_flags, end_date)
    if json_mode:
        emit_progress(4, 'Document Register', status='done', data={
            'doc_count': len(register),
            'docs': [{'doc_id': d['doc_id'], 'title': d['title'],
                       'assigned_to': d['assigned_to'], 'planned_date': d.get('planned_date')}
                      for d in register],
        })

    # ── Step 5: Show thinking + ask for confirmation ──────────────────────────
    if not json_mode:
        print_thinking(summary, project_type, ptype_data, fired_flags, register, end_date_str)
        confirmed = ask_confirmation(json_mode=False)
        if not confirmed:
            print('\n  Plan cancelled. No files written.')
            return None
    else:
        confirmed = True

    # ── Step 6: Build task assignments + write context ────────────────────────
    if json_mode:
        emit_progress(5, 'Building task assignments')
    task_assignments = build_task_assignments(summary, register)

    context = {
        'generated_at':      datetime.now().isoformat(),
        'project_end_date':  end_date_str,
        'project_type':      project_type,
        'project_summary':   summary,
        'document_register': register,
        'task_assignments':  task_assignments,
        'risk_flags_fired':  [
            {'condition': f['condition'], 'warning': f['warning'],
             'trigger_reason': f['trigger_reason']}
            for f in fired_flags
        ],
    }

    # ── Step 7: Generate handoff brief ────────────────────────────────────────
    if not json_mode:
        print('\nGenerating handoff brief...')
    if json_mode:
        emit_progress(6, 'Generating handoff brief (Gemini)')
    brief = generate_handoff_brief(context, client)
    context['handoff_brief'] = brief
    if json_mode:
        emit_progress(6, 'Handoff Brief', status='done')

    if not json_mode:
        print(f'\n{"="*64}')
        print('  PM AGENT — HANDOFF BRIEF')
        print(f'{"="*64}\n')
        print(brief)
        print()

    # ── Step 8: Generate PM-owned deliverable documents ───────────────────────
    pm_docs = [d for d in register if d.get('assigned_to') == 'PM']
    if not json_mode:
        print(f'\nGenerating PM deliverables ({len(pm_docs)} documents)...')
    if json_mode:
        emit_progress(7, f'Generating PM deliverables (0/{len(pm_docs)})')

    deliverables = generate_pm_deliverables(register, context, collection, client, json_mode=json_mode)
    context['deliverables'] = deliverables

    if json_mode:
        emit_progress(7, 'PM Deliverables', status='done', data={
            'count': len([v for v in deliverables.values() if not v.startswith('[Generation failed')]),
            'doc_ids': list(deliverables.keys()),
        })

    return context


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='PM Agent v2 — thinking loop + human confirmation'
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('--demo',     action='store_true', help='Run with built-in sample SOW')
    group.add_argument('--sow',      type=str,            help='Path to SOW text file')
    group.add_argument('--sow-text', type=str,            help='SOW as a string')
    parser.add_argument('--end-date', default='2026-06-01', help='Project end date YYYY-MM-DD')
    parser.add_argument('--save',     type=str,
                        default=str(AGENT_DIR / 'project_context.json'),
                        help='Where to save project_context.json')
    parser.add_argument('--json',     action='store_true',
                        help='Output raw JSON only (skips confirmation)')
    args = parser.parse_args()

    sow = SAMPLE_SOW if args.demo else (
        read_sow_file(args.sow) if args.sow else args.sow_text
    )

    ctx = run_pm_agent(sow, args.end_date, json_mode=args.json)

    if ctx is None:
        sys.exit(0)

    if args.json:
        print(json.dumps(ctx, indent=2, default=str))
    else:
        out = Path(args.save)
        out.write_text(json.dumps(ctx, indent=2, default=str))
        print(f'\n  ✓ project_context.json saved → {out}')

        brief_out = AGENT_DIR / 'handoff_brief.md'
        brief_out.write_text(ctx.get('handoff_brief', ''))
        print(f'  ✓ handoff_brief.md saved     → {brief_out}')

        deliverables = ctx.get('deliverables', {})
        ok_count = len([v for v in deliverables.values() if not v.startswith('[Generation failed')])
        print(f'  ✓ PM deliverables saved      → {DELIVERABLES_DIR}/ ({ok_count} files)')
        print('  Ready to hand off to Process Agent and Mechanical Agent.\n')
