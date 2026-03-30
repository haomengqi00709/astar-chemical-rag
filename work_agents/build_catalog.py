"""
Catalog Builder — generates catalog.json from real project documents.

The catalog answers one question:
  "A SOW came in — what deliverables need to be produced, by whom, in what order?"

Sources:
  - Daniel's filled Document Register result  → project type → document mapping
  - Individual work_agents documents          → what each document is for

Usage:
    python work_agents/build_catalog.py
    python work_agents/build_catalog.py --output work_agents/catalog.json
"""

import argparse
import json
import os
import re
import zipfile
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from google import genai
from google.genai import types

load_dotenv()

WORK_AGENTS = Path(__file__).parent
GENERATION_MODEL = 'gemini-2.5-flash'


# ─────────────────────────────────────────────────────────────────────────────
# Document readers
# ─────────────────────────────────────────────────────────────────────────────

def read_docx(path: Path) -> str:
    try:
        z = zipfile.ZipFile(path)
        xml = z.read('word/document.xml').decode('utf-8', errors='ignore')
        text = re.sub(r'<[^>]+>', '', xml)
        return re.sub(r'\s+', ' ', text).strip()[:3000]
    except Exception as e:
        return f'[Could not read {path.name}: {e}]'


def read_xlsx_as_text(path: Path) -> str:
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        lines = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            lines.append(f'[Sheet: {sheet}]')
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    lines.append('  ' + ' | '.join(cells))
        return '\n'.join(lines)[:3000]
    except Exception as e:
        return f'[Could not read {path.name}: {e}]'


def read_csv(path: Path) -> str:
    try:
        return path.read_text(encoding='utf-8', errors='ignore')[:2000]
    except Exception as e:
        return f'[Could not read {path.name}: {e}]'


# ─────────────────────────────────────────────────────────────────────────────
# Collect source documents
# ─────────────────────────────────────────────────────────────────────────────

def collect_sources() -> dict:
    """Read all source documents and return as a dict of name → text."""
    sources = {}

    # Daniel's filled register result — the ground truth for workflow mapping
    result_path = WORK_AGENTS / 'Daniel - Project Manager ' / 'Project Document Register- Result.xlsx'
    sources['filled_register_example'] = read_xlsx_as_text(result_path)

    # Process engineer documents
    sources['1-PRC-0001_process_design_criteria'] = read_docx(
        WORK_AGENTS / 'Aria - Process Engineer' / '1-PRC-0001-R0 Process Design Criteria.docx'
    )

    # Mechanical engineer documents
    sources['4-SPC-0002_centrifugal_pump_spec'] = read_docx(
        WORK_AGENTS / 'Hunter - Mechnical Engineer' / '4-SPC-0002-R1 Horizontal Centrifugal Pump.docx'
    )
    # .doc file — convert via LibreOffice then read, or extract raw XML
    doc_path = WORK_AGENTS / 'Hunter - Mechnical Engineer' / '4-SPC-0001-R1 General Equipment Requirements.doc'
    import subprocess, tempfile
    try:
        with tempfile.TemporaryDirectory() as tmp:
            subprocess.run(
                ['libreoffice', '--headless', '--convert-to', 'docx', '--outdir', tmp, str(doc_path)],
                capture_output=True, timeout=30
            )
            converted = Path(tmp) / '4-SPC-0001-R1 General Equipment Requirements.docx'
            sources['4-SPC-0001_general_equipment_requirements'] = read_docx(converted)
    except Exception as e:
        sources['4-SPC-0001_general_equipment_requirements'] = f'[Could not convert .doc: {e}]'
    sources['4-PRC-0007_pump_calculation_procedure'] = read_docx(
        WORK_AGENTS / 'Hunter - Mechnical Engineer' / 'Achieve' / '4-PRC-0007-R1 Pump Calculation Procedure.docx'
    )
    sources['pump_calculation_template'] = read_csv(
        WORK_AGENTS / 'Hunter - Mechnical Engineer' / 'Pump_Calculation_Template.csv'
    )

    return sources


# ─────────────────────────────────────────────────────────────────────────────
# Gemini catalog generation
# ─────────────────────────────────────────────────────────────────────────────

def build_catalog_with_gemini(sources: dict, client: genai.Client) -> dict:
    """Feed all source documents to Gemini and ask it to build the catalog."""

    sources_text = '\n\n'.join(
        f'=== SOURCE: {name} ===\n{content}'
        for name, content in sources.items()
    )

    prompt = f"""You are building a workflow catalog for an engineering firm's AI agent system.

The catalog answers one question only:
"When a Scope of Work (SOW) comes in describing a project, what deliverables
need to be produced, by whom, and in what order?"

You have been given the following source documents to derive the catalog from:

{sources_text}

Based ONLY on these source documents (not general knowledge), generate a catalog JSON.

The catalog must have this exact structure:
{{
  "version": "1.0",
  "built_from": ["list of source document names you used"],
  "project_types": {{
    "centrifugal_pump": {{
      "description": "What kind of project this is",
      "triggers": ["keywords in a SOW that identify this project type"],
      "disciplines_involved": [0, 1, 4],
      "agents": ["PM", "Process Engineer", "Mechanical Engineer"],
      "required_docs": [
        {{
          "doc_id": "e.g. 1-CAL-6510",
          "type": "CAL/DST/SPC/LST/TAB/PRC/DAT",
          "title": "document title",
          "discipline": 0,
          "discipline_name": "e.g. Process Technology",
          "assigned_to": "PM / Process Engineer / Mechanical Engineer",
          "reason": "Why this document is needed — derived from the source documents",
          "inputs_from": [],
          "outputs_to": [],
          "depends_on": [],
          "weeks_before_end": 4
        }}
      ],
      "risk_flags": [
        {{
          "condition": "short name e.g. corrosive_fluid",
          "description": "what triggers this flag",
          "sow_keywords": ["keywords that trigger this flag"],
          "sow_thresholds": {{}},
          "additional_docs": ["doc_ids to add when flagged"],
          "warning": "what the PM agent should tell the user",
          "rag_query": "what to query RAG for when this flag fires"
        }}
      ],
      "scheduling_notes": "plain text description of the scheduling logic"
    }}
  }}
}}

Rules:
- Base EVERY decision on the source documents provided — do not invent documents
- The reason field for each doc must reference which source document justifies its inclusion
- Risk flags must be grounded in what the source documents actually say about special conditions
- The depends_on chain must reflect the actual workflow sequence shown in the filled register
- IMPORTANT: 4-DST-XXXX (Pump Data Sheet template) is always owned by Mechanical Engineer.
  In 4-PRC-0007, "PDS" means Process Data Sheet (an input document), NOT the 4-DST-XXXX template.
- IMPORTANT: Always include corrosive_fluid and high_pressure as risk flags — derive their
  keywords and RAG queries from what the source documents say about material selection and pressure testing.
  For corrosive_fluid, set additional_docs to ["4-SPC-0001"] since general equipment requirements
  apply when special service conditions exist.
- Return ONLY the JSON, no explanation, no markdown fences"""

    response = client.models.generate_content(
        model=GENERATION_MODEL,
        contents=prompt,
        config=types.GenerateContentConfig(temperature=0.1),
    )

    text = response.text.strip()
    if text.startswith('```'):
        parts = text.split('```')
        text = parts[1]
        if text.startswith('json'):
            text = text[4:]
    return json.loads(text.strip())


# ─────────────────────────────────────────────────────────────────────────────
# Corrections — applied after Gemini output to fix known misassignments
# ─────────────────────────────────────────────────────────────────────────────

# Keyed by doc_id — fields listed here override whatever Gemini produced
CORRECTIONS = {
    '4-DST-XXXX': {
        'assigned_to': 'Mechanical Engineer',
        'discipline': 4,
        'discipline_name': 'Equipment',
    },
    '1-PRC-0001': {
        'assigned_to': 'Process Engineer',
        'discipline': 1,
        'discipline_name': 'Process Technology',
        'weeks_before_end': 6,
    },
    '1-CAL-6510': {
        'assigned_to': 'Process Engineer',
        'discipline': 1,
        'discipline_name': 'Process Technology',
        'weeks_before_end': 4,
    },
    '4-CAL-0001': {
        'assigned_to': 'Mechanical Engineer',
        'discipline': 4,
        'discipline_name': 'Equipment',
        'weeks_before_end': 3,
    },
    '4-SPC-0002': {
        'assigned_to': 'Mechanical Engineer',
        'discipline': 4,
        'discipline_name': 'Equipment',
        'weeks_before_end': 2,
    },
    '4-TAB-XXXX': {
        'assigned_to': 'Mechanical Engineer',
        'discipline': 4,
        'discipline_name': 'Equipment',
        'weeks_before_end': 1,
    },
}


def apply_corrections(catalog: dict) -> dict:
    """Override known misassignments in Gemini output."""
    for ptype in catalog.get('project_types', {}).values():
        for doc in ptype.get('required_docs', []):
            if doc['doc_id'] in CORRECTIONS:
                doc.update(CORRECTIONS[doc['doc_id']])
    return catalog


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Build catalog.json from project documents')
    parser.add_argument('--output', default=str(WORK_AGENTS / 'catalog.json'),
                        help='Output path for catalog.json')
    args = parser.parse_args()

    api_key = os.environ.get('GOOGLE_API_KEY')
    if not api_key:
        raise RuntimeError('Set GOOGLE_API_KEY environment variable.')

    client = genai.Client(api_key=api_key)

    print('[1/3] Reading source documents...')
    sources = collect_sources()
    for name, content in sources.items():
        status = 'OK' if not content.startswith('[Could not read') else 'MISSING'
        print(f'      {status}  {name}')

    print('[2/3] Sending to Gemini to build catalog...')
    catalog = build_catalog_with_gemini(sources, client)
    catalog = apply_corrections(catalog)

    print('[3/3] Saving catalog...')
    out = Path(args.output)
    out.write_text(json.dumps(catalog, indent=2))
    print(f'      Saved → {out}')

    # Print summary
    print()
    for ptype, pdata in catalog.get('project_types', {}).items():
        docs = pdata.get('required_docs', [])
        flags = pdata.get('risk_flags', [])
        print(f'Project type: {ptype}')
        print(f'  {len(docs)} required documents:')
        for d in docs:
            print(f'    {d["doc_id"]:<18} {d["assigned_to"]:<22} {d["title"]}')
        print(f'  {len(flags)} risk flags:')
        for f in flags:
            print(f'    [{f["condition"]}] {f["warning"][:80]}')


if __name__ == '__main__':
    main()
