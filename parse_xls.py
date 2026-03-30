"""
XLS/XLSX parser for RAG compliance system.

Parses all Excel files found in List, Procedure, Datasheet, and Specification
subfolders across engineering disciplines.
Outputs structured text chunks with metadata, and flags rows that need LLM review.

Each chunk carries a source_folder field (List / Procedure / Datasheet / Specification)
enabling demo filtering:
  - Baseline mode:  source_folder in [List, Procedure]
  - Enhanced mode:  all 4 folders (no filter)

Usage:
    python parse_xls.py <root_dir>
    python parse_xls.py .

Dependencies:
    pip install openpyxl xlrd
    (openpyxl handles .xlsx; xlrd handles legacy .xls)
"""

import re
import json
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Filename metadata detection
# ---------------------------------------------------------------------------

def detect_discipline(filename: str) -> int | None:
    match = re.match(r'^(\d+)-', filename)
    return int(match.group(1)) if match else None

def detect_doc_id(filename: str) -> str:
    match = re.match(r'^(\d+-[A-Z]+-[\d\-]+)', filename)
    return match.group(1).rstrip('-') if match else filename

def detect_doc_type(filename: str) -> str:
    for t in ('LST', 'PRC', 'DST', 'CAL', 'SPC', 'QDS'):
        if f'-{t}-' in filename:
            return t
    return 'UNKNOWN'

def detect_revision(filename: str) -> str | None:
    match = re.search(r'-R(\d+)', filename, re.IGNORECASE)
    return f'R{match.group(1)}' if match else None

def detect_is_template(filename: str) -> bool:
    return bool(re.search(r'\[?[Tt]emplate\]?|\([Tt]emplate\)', filename))

ALLOWED_FOLDERS = {'List', 'Procedure', 'Procedures'}
# Datasheet and Specification excluded for now — may be added in future

DISCIPLINE_NAMES = {
    0: 'Administration',
    1: 'Process Technology',
    2: 'Civil & Structural',
    3: 'Mechanical',
    4: 'Equipment',
    5: 'Piping & Layout',
    6: 'Insulation',
    7: 'Coatings',
    8: 'Instrumentation & Control',
    9: 'Electrical',
}

def detect_source_folder(file_path: Path) -> str:
    """Walk up the path and return the first ancestor that matches an allowed folder."""
    for part in reversed(file_path.parts):
        if part in ('Procedure', 'Procedures'):
            return 'Procedure'
        if part in ('Specification', 'Specifications'):
            return 'Specification'
        if part in ('List', 'Datasheet'):
            return part
    return 'Unknown'

def detect_discipline_name(discipline: int | None) -> str:
    return DISCIPLINE_NAMES.get(discipline, 'Unknown')


# ---------------------------------------------------------------------------
# Discipline-aware row-to-attribute conversion
# ---------------------------------------------------------------------------

def _get(row: dict, *keys, fallback: str = '') -> str:
    """Try multiple possible column names and return the first match.
    Falls back to partial key matching (key is a substring of the row key)
    to handle cases where multi-row headers merge column names with data values.
    """
    for key in keys:
        # Exact match
        val = row.get(key)
        if val is not None and str(val).strip():
            return str(val).strip()
    for key in keys:
        # Partial match — key is contained in the row key (case-insensitive)
        key_lower = key.lower()
        for row_key, row_val in row.items():
            if key_lower in row_key.lower() and row_val is not None and str(row_val).strip():
                return str(row_val).strip()
    return fallback


def row_to_text(row: dict, discipline: int | None, doc_id: str, sheet_name: str) -> str:
    """Convert a data row dict into a natural language attribute sentence."""

    # --- Discipline 1: Fluid Codes (1-LST-0002) ---
    # Actual headers: Our Scope: B/C/D | Selection | Fluid | Code | Comments | Technology
    if '1-LST-0002' in doc_id:
        return (
            f"Fluid code [{_get(row, 'Code', 'Fluid Code')}] "
            f"represents fluid [{_get(row, 'Fluid', 'Fluid Name')}], "
            f"scope [{_get(row, 'Our Scope: B/C/D', 'Selection')}], "
            f"technology [{_get(row, 'Technology')}], "
            f"comments [{_get(row, 'Comments')}]."
        )

    # --- Discipline 5: Line List (5-LST-0001) ---
    # Actual merged headers (3 rows combined with /):
    # Line Number / Pipe / Size | Insul. / Fluid / Code | P&ID / Mat'l / Code |
    # Temperature / Operating / Min. | Pressure / Design | etc.
    if '5-LST-0001' in doc_id:
        return (
            f"Line [{_get(row, 'Line Number / Pipe / Size', 'Line Number', 'Line No')}] "
            f"carries fluid code [{_get(row, 'Insul. / Fluid / Code', 'Fluid', 'Fluid Code')}], "
            f"material class [{_get(row, 'P&ID / Mat\'l / Code', 'Material', 'Mat')}], "
            f"design pressure [{_get(row, 'Pressure / Design', 'Pressure', 'Design Pressure')}] kPag, "
            f"operating temperature [{_get(row, 'Temperature / Operating / Min.', 'Temperature')}] degC, "
            f"insulation [{_get(row, 'Insulation', 'Insul.')}]."
        )

    # --- Discipline 5: Tie Point List (5-LST-0002) ---
    if '5-LST-0002' in doc_id:
        return (
            f"Tie point [{_get(row, 'Tie Point', 'TP No', 'TP')}] "
            f"on line [{_get(row, 'Line No', 'Line')}], "
            f"size [{_get(row, 'Size')}], "
            f"pressure rating [{_get(row, 'Press', 'Pressure', 'Rating')}] kPag, "
            f"location [{_get(row, 'Location', 'Description')}]."
        )

    # --- Discipline 5: Piping Service List (5-LST-0003-x) ---
    # Actual headers: Service Code | Fluid | Fluid Composition |
    # Material Description / Application | ASME B31.3 Fluid Class |
    # Pressure Test Method | Flange Guard | Corr. Allow. (mm)
    if '5-LST-0003' in doc_id:
        return (
            f"Piping service [{_get(row, 'Service Code', 'Code')}] "
            f"fluid [{_get(row, 'Fluid', 'Fluid Name')}], "
            f"composition [{_get(row, 'Fluid Composition', 'Composition')}], "
            f"material [{_get(row, 'Material Description / Application', 'Material Description', 'Material')}], "
            f"ASME fluid class [{_get(row, 'ASME B31.3 Fluid Class', 'Fluid Class')}], "
            f"pressure test [{_get(row, 'Pressure Test Method', 'Test Method')}], "
            f"corrosion allowance [{_get(row, 'Corr. Allow.\n(mm)', 'Corr. Allow. (mm)', 'Corr Allow')}] mm, "
            f"plant type [{sheet_name}]."
        )

    # --- Discipline 8: Alarm Interlocks & Setpoints (8-LST-0103) ---
    if '8-LST-0103' in doc_id:
        return (
            f"Tag [{_get(row, 'Tag', 'Instrument Tag')}] "
            f"has alarm type [{_get(row, 'Alarm Type', 'Alarm')}] "
            f"at setpoint [{_get(row, 'Setpoint', 'Set Point')}] [{_get(row, 'Unit', 'Units')}], "
            f"interlock action [{_get(row, 'Action', 'Interlock Action')}] "
            f"at [{_get(row, 'Interlock Value', 'Trip Value', 'Trip')}] [{_get(row, 'Unit', 'Units')}]."
        )

    # --- Discipline 8: Instrument Process Data Lists (8-LST-0110 to 0121) ---
    if discipline == 8 and re.search(r'8-LST-01', doc_id):
        return (
            f"Instrument [{_get(row, 'Tag', 'Instrument Tag')}] "
            f"type [{sheet_name}] "
            f"measures [{_get(row, 'Process Variable', 'Variable', 'Measured Variable')}] "
            f"on [{_get(row, 'Line No', 'Equipment', 'Line/Equip')}], "
            f"range [{_get(row, 'Range')}] [{_get(row, 'Unit', 'Units')}], "
            f"design pressure [{_get(row, 'Press', 'Design Pressure')}] kPag, "
            f"design temperature [{_get(row, 'Temp', 'Design Temp')}] degC, "
            f"fluid code [{_get(row, 'Fluid', 'Fluid Code')}]."
        )

    # --- Discipline 9: Electrical Checkout Lists (9-LST-07xx) ---
    if discipline == 9:
        return (
            f"System [{_get(row, 'System', fallback=sheet_name)}], "
            f"check item [{_get(row, 'Item No', 'Item', 'No')}]: "
            f"[{_get(row, 'Requirement', 'Description', 'Check')}]. "
            f"Acceptance criteria: [{_get(row, 'Criteria', 'Acceptance', 'Pass Criteria')}]."
        )

    # --- Generic fallback: key: value pairs ---
    parts = [f"{k}: {v}" for k, v in row.items() if v and str(v).strip()]
    return ' | '.join(parts)


# ---------------------------------------------------------------------------
# Quality check — flags rows that may need LLM review
# ---------------------------------------------------------------------------

def quality_check(text: str, row: dict) -> list[str]:
    issues = []

    # Too many empty fields in the row
    total = len(row)
    empty = sum(1 for v in row.values() if not v or str(v).strip() == '')
    if total > 0 and empty / total > 0.5:
        issues.append(f'{empty}/{total} fields are empty')

    # Converted text too short to be useful
    if len(text.strip()) < 20:
        issues.append('converted text too short')

    # Too many empty bracket pairs in the output — key fields had no value
    empty_brackets = re.findall(r'\[\s*\]', text)
    if len(empty_brackets) > 3:
        issues.append(f'{len(empty_brackets)} key fields resolved to empty')

    return issues


# ---------------------------------------------------------------------------
# Sheet parser
# ---------------------------------------------------------------------------

def parse_sheet(ws, discipline: int | None, doc_id: str, sheet_name: str,
                is_template: bool, revision: str | None,
                source_folder: str = '') -> tuple[list, list]:
    """Parse one Excel sheet into chunks. Returns (chunks, needs_llm_review)."""
    chunks = []
    needs_llm = []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return chunks, needs_llm

    # Skip cover/signature sheets — first non-empty row has '|' AND numeric column markers
    first_non_empty = next(
        (row for row in rows if any(c is not None and str(c).strip() for c in row)), None
    )
    if first_non_empty and str(first_non_empty[0]).strip() == '|':
        numeric_after = [c for c in first_non_empty[1:]
                         if c is not None and str(c).strip().replace('.', '').isdigit()]
        if len(numeric_after) >= 3:
            return chunks, needs_llm

    # Find header row: first row with 4+ non-empty, non-numeric cells
    # (threshold of 4 skips title rows like "Client Name | Project Title | ...")
    header_idx = None
    for i, row in enumerate(rows):
        non_empty = [c for c in row if c is not None and str(c).strip()
                     and not str(c).strip().replace('.', '').isdigit()]
        if len(non_empty) >= 4:
            header_idx = i
            break

    if header_idx is None:
        return chunks, needs_llm

    # Build headers — merge up to 2 consecutive header-looking rows (handles merged cells)
    # Only merge if avg cell length > 5 chars — prevents data rows like ['A','WA','C']
    # from being absorbed into headers
    header_parts = [rows[header_idx]]
    for offset in (1, 2):
        next_idx = header_idx + offset
        if next_idx >= len(rows):
            break
        candidate = rows[next_idx]
        non_empty_vals = [str(c).strip() for c in candidate
                          if c is not None and str(c).strip()
                          and not str(c).strip().replace('.', '').isdigit()]
        if len(non_empty_vals) < 2:
            break
        avg_len = sum(len(v) for v in non_empty_vals) / len(non_empty_vals)
        if avg_len <= 5:
            break
        header_parts.append(candidate)

    # Combine multi-row headers: join non-empty values per column with '/'
    num_cols = max(len(r) for r in header_parts)
    headers = []
    for col in range(num_cols):
        parts = [str(r[col]).strip() for r in header_parts
                 if col < len(r) and r[col] is not None and str(r[col]).strip()]
        headers.append(' / '.join(parts) if parts else '')

    data_start = header_idx + len(header_parts)

    for row_num, raw_row in enumerate(rows[data_start:], start=data_start + 1):
        # Skip fully empty rows
        if all(c is None or str(c).strip() == '' for c in raw_row):
            continue

        row_dict = {
            headers[i]: raw_row[i] if i < len(raw_row) else None
            for i in range(len(headers))
            if headers[i]  # skip columns with no header
        }

        text = row_to_text(row_dict, discipline, doc_id, sheet_name)
        issues = quality_check(text, row_dict)

        chunk = {
            'text': text,
            'metadata': {
                'doc_id': doc_id,
                'doc_type': detect_doc_type(doc_id),
                'discipline': discipline,
                'discipline_name': detect_discipline_name(discipline),
                'source_folder': source_folder,
                'sheet': sheet_name,
                'row': row_num,
                'revision': revision,
                'is_template': is_template,
                'needs_llm_review': bool(issues),
                'review_reasons': issues if issues else None,
                # Keep raw row for debugging and LLM fallback input
                'raw': {k: str(v) for k, v in row_dict.items() if v is not None},
            }
        }

        chunks.append(chunk)
        if issues:
            needs_llm.append(chunk)

    return chunks, needs_llm


# ---------------------------------------------------------------------------
# File-level parser
# ---------------------------------------------------------------------------

def parse_xls(file_path: str | Path) -> tuple[list, list]:
    """Parse an XLS/XLSX file. Returns (all_chunks, needs_llm_review)."""
    path = Path(file_path)
    filename = path.name

    discipline = detect_discipline(filename)
    doc_id = detect_doc_id(filename)
    revision = detect_revision(filename)
    is_template = detect_is_template(filename)
    source_folder = detect_source_folder(path)

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        # openpyxl can't open legacy .xls (pre-2007) — needs xlrd
        try:
            import xlrd
            return _parse_xls_legacy(path, discipline, doc_id, revision, is_template, source_folder)
        except ImportError:
            raise RuntimeError(
                f"Cannot open {filename}: install xlrd for legacy .xls support  →  pip install xlrd"
            )

    all_chunks, all_needs_llm = [], []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        chunks, needs_llm = parse_sheet(
            ws, discipline, doc_id, sheet_name, is_template, revision, source_folder
        )
        all_chunks.extend(chunks)
        all_needs_llm.extend(needs_llm)

    return all_chunks, all_needs_llm


def _parse_xls_legacy(path: Path, discipline, doc_id, revision, is_template, source_folder=''):
    """Fallback parser for legacy .xls files using xlrd."""
    import xlrd

    wb = xlrd.open_workbook(str(path))
    all_chunks, all_needs_llm = [], []

    for sheet in wb.sheets():
        rows = [sheet.row_values(i) for i in range(sheet.nrows)]
        if not rows:
            continue

        # Skip cover/signature sheets — first non-empty row has '|' AND numeric column markers
        first_non_empty = next((r for r in rows if any(c and str(c).strip() for c in r)), None)
        if first_non_empty and str(first_non_empty[0]).strip() == '|':
            numeric_after = [c for c in first_non_empty[1:]
                             if c and str(c).strip().replace('.', '').isdigit()]
            if len(numeric_after) >= 3:
                continue

        # Find header row: first row with 4+ non-empty, non-numeric cells
        # (threshold of 4 skips title rows like "Client Name | Project Title | ...")
        header_idx = None
        for i, row in enumerate(rows):
            non_empty = [c for c in row if c and str(c).strip()
                         and not str(c).strip().replace('.', '').isdigit()]
            if len(non_empty) >= 4:
                header_idx = i
                break
        if header_idx is None:
            continue

        # Build headers — merge up to 2 consecutive header-looking rows
        # Only merge if avg cell length > 5 chars — prevents short data codes from being absorbed
        header_parts = [rows[header_idx]]
        for offset in (1, 2):
            next_idx = header_idx + offset
            if next_idx >= len(rows):
                break
            candidate = rows[next_idx]
            non_empty_vals = [str(c).strip() for c in candidate
                              if c and str(c).strip()
                              and not str(c).strip().replace('.', '').isdigit()]
            if len(non_empty_vals) < 2:
                break
            avg_len = sum(len(v) for v in non_empty_vals) / len(non_empty_vals)
            if avg_len <= 5:
                break
            header_parts.append(candidate)

        num_cols = max(len(r) for r in header_parts)
        headers = []
        for col in range(num_cols):
            parts = [str(r[col]).strip() for r in header_parts
                     if col < len(r) and r[col] and str(r[col]).strip()]
            headers.append(' / '.join(parts) if parts else '')

        data_start = header_idx + len(header_parts)

        for row_num, raw_row in enumerate(rows[data_start:], start=data_start + 1):
            if all(not c or str(c).strip() == '' for c in raw_row):
                continue

            row_dict = {
                headers[i]: raw_row[i] if i < len(raw_row) else None
                for i in range(len(headers))
                if headers[i]
            }

            text = row_to_text(row_dict, discipline, doc_id, sheet.name)
            issues = quality_check(text, row_dict)

            chunk = {
                'text': text,
                'metadata': {
                    'doc_id': doc_id,
                    'doc_type': detect_doc_type(doc_id),
                    'discipline': discipline,
                    'source_folder': source_folder,
                    'sheet': sheet.name,
                    'row': row_num,
                    'revision': revision,
                    'is_template': is_template,
                    'needs_llm_review': bool(issues),
                    'review_reasons': issues if issues else None,
                    'raw': {k: str(v) for k, v in row_dict.items() if v is not None},
                }
            }
            all_chunks.append(chunk)
            if issues:
                all_needs_llm.append(chunk)

    return all_chunks, all_needs_llm


# ---------------------------------------------------------------------------
# Batch runner
# ---------------------------------------------------------------------------

def parse_all_xls(root_dir: str | Path) -> tuple[list, list, list]:
    """
    Walk root_dir and parse all XLS/XLSX files inside List, Procedure, Datasheet,
    and Specification subfolders only. All other subfolders are ignored.

    Each chunk carries source_folder for demo filtering:
      Baseline:  source_folder in [List, Procedure]
      Enhanced:  all 4 folders

    Returns (all_chunks, needs_llm_review, report).
    """
    root = Path(root_dir)
    all_chunks, all_needs_llm, report = [], [], []

    xls_files = sorted(
        f for f in root.rglob('*')
        if f.suffix.lower() in ('.xls', '.xlsx')
        and any(part in ALLOWED_FOLDERS for part in f.parts)
    )

    for xls_file in xls_files:
        rel = xls_file.relative_to(root)
        print(f"Parsing: {rel}")
        try:
            chunks, needs_llm = parse_xls(xls_file)
            all_chunks.extend(chunks)
            all_needs_llm.extend(needs_llm)
            report.append({
                'file': str(rel),
                'source_folder': detect_source_folder(xls_file),
                'chunks': len(chunks),
                'needs_llm_review': len(needs_llm),
                'status': 'ok',
            })
            print(f"  → {len(chunks)} chunks, {len(needs_llm)} flagged for LLM review")
        except Exception as e:
            print(f"  ERROR: {e}")
            report.append({'file': str(rel), 'status': 'error', 'error': str(e)})

    return all_chunks, all_needs_llm, report


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    import sys

    root = sys.argv[1] if len(sys.argv) > 1 else '.'

    chunks, needs_llm, report = parse_all_xls(root)

    # Always write output next to the script, not inside the data folder
    out_dir = Path(__file__).parent

    # Save all chunks
    chunks_path = out_dir / 'parsed_chunks.json'
    with open(chunks_path, 'w') as f:
        json.dump(chunks, f, indent=2, default=str)
    print(f"\nSaved {len(chunks)} chunks → {chunks_path}")

    # Save LLM review queue (only if there are items)
    if needs_llm:
        llm_path = out_dir / 'llm_review_queue.json'
        with open(llm_path, 'w') as f:
            json.dump(needs_llm, f, indent=2, default=str)
        print(f"Saved {len(needs_llm)} items for LLM review → {llm_path}")

    # Print summary report
    print("\n--- Parse Report ---")
    total_chunks = sum(r.get('chunks', 0) for r in report)
    total_llm = sum(r.get('needs_llm_review', 0) for r in report)
    errors = [r for r in report if r.get('status') == 'error']

    for entry in report:
        status = entry.get('status', '')
        if status == 'error':
            print(f"  ERROR  {entry['file']}: {entry['error']}")
        else:
            llm_flag = f"  ({entry['needs_llm_review']} need LLM review)" if entry['needs_llm_review'] else ''
            print(f"  OK     {entry['file']} — {entry['chunks']} chunks{llm_flag}")

    print(f"\nTotal: {total_chunks} chunks, {total_llm} flagged, {len(errors)} errors")
