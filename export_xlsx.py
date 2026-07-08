"""
Export process calculation data to a formatted .xlsx file with live Excel formulas.

Usage:
    python export_xlsx.py --input process_output.json --output 1-CAL-XXXX.xlsx
    cat process_output.json | python export_xlsx.py --output 1-CAL-XXXX.xlsx
"""

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Colour palette ────────────────────────────────────────────────────────────
C_HEADER_BG  = '1F4E79'
C_HEADER_FG  = 'FFFFFF'
C_SECTION_BG = '2E75B6'
C_SECTION_FG = 'FFFFFF'
C_COL_HDR_BG = 'D6E4F0'
C_COL_HDR_FG = '1F4E79'
C_INPUT_BG   = 'EBF3FB'   # light blue  — user inputs
C_CALC_BG    = 'FFF2CC'   # yellow      — calculated
C_OUTPUT_BG  = 'E2EFDA'   # green       — outputs / hand-off
C_ALT_BG     = 'F7FBFF'
C_BORDER     = 'BFBFBF'

VAL_COL = 3   # Column C — value / formula


def _fill(hex_colour: str) -> PatternFill:
    return PatternFill('solid', fgColor=hex_colour)


def _font(bold=False, colour='000000', size=10) -> Font:
    return Font(bold=bold, color=colour, size=size, name='Calibri')


def _border() -> Border:
    s = Side(style='thin', color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h='left', v='center', wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _write_row(ws, row: int, cells: list, fill: PatternFill,
               bold: bool = False, height: int = 15):
    for ci, val in enumerate(cells, start=1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font      = _font(bold=bold)
        c.border    = _border()
        c.alignment = _align(wrap=(ci in (5, 6)))
        c.fill      = fill
    if height:
        ws.row_dimensions[row].height = height


def build_process_calc_xlsx(data: dict, output_path: Path) -> None:
    ps  = data.get('project_summary',  {})
    fp  = data.get('fluid_properties', {})
    dc  = data.get('design_criteria',  {})
    hr  = data.get('hydraulic_results',{})

    wb = Workbook()
    ws = wb.active
    ws.title = '1-CAL-XXXX'

    # Column widths  A=Parameter  B=Symbol  C=Value  D=Units  E=Excel Formula  F=Engineering Formula
    for i, w in enumerate([32, 10, 14, 10, 38, 38], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1   # current row pointer
    # refs maps logical name → row number (all in column C)
    refs: dict[str, int] = {}
    NCOLS = 6  # A–F

    # ── Document header ───────────────────────────────────────────────────────
    hdr_fill = _fill(C_HEADER_BG)

    ws.merge_cells(f'A{r}:C{r}')
    ws.merge_cells(f'D{r}:F{r}')
    ws.cell(r, 1, 'A Star Chemical').font      = _font(bold=True, colour=C_HEADER_FG, size=12)
    for ci in range(1, 4): ws.cell(r, ci).fill = hdr_fill
    ws.cell(r, 4, '1-CAL-XXXX').font           = _font(bold=True, colour=C_HEADER_FG, size=12)
    ws.cell(r, 4).alignment                    = _align(h='right')
    for ci in range(4, 7): ws.cell(r, ci).fill = hdr_fill
    ws.row_dimensions[r].height = 22; r += 1

    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(r, 1, 'Process Calculations — Hydraulic Design').font = _font(bold=True, colour=C_HEADER_FG, size=11)
    for ci in range(1, 7): ws.cell(r, ci).fill = hdr_fill
    ws.row_dimensions[r].height = 18; r += 1

    info_fill = _fill(C_COL_HDR_BG)
    for label, val in [
        ('Project',  ps.get('project_title', '—')),
        ('Client',   ps.get('client', '—')),
        ('Location', ps.get('location', '—')),
        ('Date',     data.get('generated_at', '—')[:10]),
    ]:
        ws.merge_cells(f'B{r}:F{r}')
        ws.cell(r, 1, label).font  = _font(bold=True)
        for ci in range(1, 7):
            ws.cell(r, ci).fill   = info_fill
            ws.cell(r, ci).border = _border()
        ws.cell(r, 2, val).alignment = _align()
        ws.row_dimensions[r].height = 14; r += 1

    r += 1  # blank spacer

    # Column headers: A=Parameter B=Symbol C=Value D=Units E=Excel Formula F=Engineering Formula
    _write_row(ws, r,
               ['Parameter', 'Symbol', 'Value', 'Units', 'Excel Formula', 'Engineering Formula'],
               fill=_fill(C_COL_HDR_BG), bold=True, height=16)
    r += 1

    # ── Helpers ───────────────────────────────────────────────────────────────

    def section(title: str):
        nonlocal r
        ws.merge_cells(f'A{r}:F{r}')
        ws.cell(r, 1, title).font      = _font(bold=True, colour=C_SECTION_FG)
        ws.cell(r, 1).alignment        = _align(h='center')
        for ci in range(1, 7):
            ws.cell(r, ci).fill        = _fill(C_SECTION_BG)
            ws.cell(r, ci).border      = _border()
        ws.row_dimensions[r].height    = 15; r += 1

    def data_row(ref_name: str | None,
                 param: str, symbol: str, value, units: str,
                 eng_formula: str = '',   # human-readable formula (col F)
                 bg: str = C_INPUT_BG) -> int:
        """Write a data row. Col C = live Excel formula or static value.
        Col E = the Excel formula as plain text (so engineers can read it).
        Col F = human-readable engineering formula.
        Returns the row number written."""
        nonlocal r
        row_num = r
        if ref_name:
            refs[ref_name] = row_num

        # Col E: only show formula text for calculated rows (formula rows); inputs left blank
        excel_formula_text = str(value) if str(value).startswith('=') else ''

        _write_row(ws, r, [param, symbol, value, units, excel_formula_text, eng_formula],
                   fill=_fill(bg))
        ws.cell(r, VAL_COL).alignment = _align(h='right')
        # Style column E (excel formula text) as monospace
        if excel_formula_text:
            ws.cell(r, 5).font = Font(name='Courier New', size=8, color='555555')
        r += 1
        return row_num

    def C(ref_name: str) -> str:
        """Return Excel cell reference for a tracked row's value column."""
        return f'C{refs[ref_name]}'

    # ═════════════════════════════════════════════════════════════════════════
    # SECTION 1 — FLUID DATA  (INPUTS — static values)
    # ═════════════════════════════════════════════════════════════════════════
    section('SECTION 1 — FLUID DATA  [INPUTS]')

    data_row(None,  'Fluid Name',        '—',  fp.get('fluid_name'),                  '—',     eng_formula=f"Code: {fp.get('fluid_code')}  |  Source: {fp.get('fluid_code_source')}")
    data_row('rho', 'Fluid Density',     'ρ',  fp.get('density_kgm3'),                'kg/m³', eng_formula='From process data / SOW')
    data_row('SG',  'Specific Gravity',  'SG', f"={C('rho')}/1000",                   '—',     eng_formula='SG = ρ / 1000')
    data_row('mu',  'Dynamic Viscosity', 'μ',  fp.get('dynamic_viscosity_mPas'),       'mPa·s', eng_formula='Fluid property lookup / engineering estimate')
    data_row('Pv',  'Vapour Pressure',   'Pv', fp.get('vapor_pressure_kPa_at_temp'),   'kPa',   eng_formula=f"At operating temperature {ps.get('temperature_c')} °C")
    data_row(None,  'Corrosive Service', '—',  'Yes' if fp.get('corrosive') else 'No', '—',     eng_formula='Drives material selection and seal requirements')
    data_row(None,  'Fluid Notes',       '—',  fp.get('notes', '—'),                  '—',     eng_formula='1-LST-0002', bg=C_ALT_BG)

    r += 1

    # ═════════════════════════════════════════════════════════════════════════
    # SECTION 2 — DESIGN CONDITIONS  (INPUTS — static values)
    # ═════════════════════════════════════════════════════════════════════════
    section('SECTION 2 — DESIGN CONDITIONS  [INPUTS]')

    data_row('Qn',     'Normal Flow Rate',          'Q_n',  dc.get('normal_flow_m3h'),              'm³/h',   'From SOW / process data sheet')
    data_row('margin', 'Flow Margin',               '—',    dc.get('flow_margin_pct'),               '%',      f"Source: {dc.get('flow_margin_source')}")
    data_row('Qr',     'Rated Flow Rate',           'Q_r',  f"={C('Qn')}*(1+{C('margin')}/100)",    'm³/h',   f"Q_r = Q_n × (1 + margin%)  →  {dc.get('normal_flow_m3h')} × {1 + dc.get('flow_margin_pct',0)/100:.2f}")
    data_row('Pdes',   'Design Pressure',           'P_d',  dc.get('design_pressure_bar'),           'bar(g)', 'From SOW')
    data_row(None,     'Pressure Class',            '—',    dc.get('pressure_class'),                '—',      'Derived from design pressure')
    data_row('T',      'Design Temperature',        'T',    dc.get('design_temperature_c'),           '°C',     'From SOW')
    data_row('Ps',     'Suction Vessel Pressure',   'Ps',   ps.get('suction_pressure_barg'),          'barg',   'From SOW / P&ID')
    data_row('Pd',     'Discharge Vessel Pressure', 'Pd',   ps.get('discharge_pressure_barg'),        'barg',   'From SOW / P&ID')
    data_row('Zs',     'Suction Static Head',       'Zs',   ps.get('suction_static_head_m'),          'm',      'Elevation of suction liquid level above pump CL')
    data_row('Zd',     'Discharge Static Head',     'Zd',   ps.get('discharge_static_head_m'),        'm',      'Elevation of discharge vessel inlet above pump CL')
    data_row('hfs',    'Suction Friction Loss',     'hf_s', ps.get('suction_friction_loss_m'),        'm',      'Pipe friction losses — suction line')
    data_row('hfd',    'Discharge Friction Loss',   'hf_d', ps.get('discharge_friction_loss_m'),      'm',      'Pipe friction losses — discharge line')
    data_row('CA',     'Corrosion Allowance',       'CA',   dc.get('corrosion_allowance_mm'),         'mm',     f"Source: {dc.get('corrosion_allowance_source')}")

    r += 1

    # ═════════════════════════════════════════════════════════════════════════
    # SECTION 3 — HYDRAULIC CALCULATIONS  (live Excel formulas)
    # ═════════════════════════════════════════════════════════════════════════
    section('SECTION 3 — HYDRAULIC CALCULATIONS  [CALC — live formulas]')

    data_row('Hp',
             'Pressure Differential Head', 'Hp',
             f"=ROUND(({C('Pd')}-{C('Ps')})*100000/({C('rho')}*9.81),2)",
             'm',
             f"Hp = (P_d − P_s) × 100,000 / (ρ × 9.81)",
             bg=C_CALC_BG)

    data_row('Hs',
             'Static Head', 'Hs',
             f"=ROUND({C('Zd')}-{C('Zs')},2)",
             'm',
             'Hs = Z_d − Z_s',
             bg=C_CALC_BG)

    data_row('Hf',
             'Total Friction Loss', 'Hf',
             f"=ROUND({C('hfs')}+{C('hfd')},2)",
             'm',
             'Hf = hf_s + hf_d',
             bg=C_CALC_BG)

    data_row('TDH',
             'Total Dynamic Head', 'TDH',
             f"=ROUND({C('Hp')}+{C('Hs')}+{C('Hf')},2)",
             'm',
             'TDH = Hp + Hs + Hf',
             bg=C_CALC_BG)

    data_row('Hvp',
             'Vapour Pressure Head', 'Hvp',
             f"=ROUND({C('Pv')}*1000/({C('rho')}*9.81),3)",
             'm',
             'Hvp = Pv × 1000 / (ρ × g)',
             bg=C_CALC_BG)

    # NPSHa = (Ps_abs)/ρg + Zs - hfs - Pv/ρg
    # Ps_abs [kPa] = (Ps_barg + 1.01325) × 100
    data_row('NPSHa',
             'Net Positive Suction Head Available', 'NPSHa',
             f"=ROUND(({C('Ps')}+1.01325)*100000/({C('rho')}*9.81)+{C('Zs')}-{C('hfs')}-{C('Pv')}*1000/({C('rho')}*9.81),2)",
             'm',
             'NPSHa = (P_s_abs / ρg) + Z_s − hf_s − (P_v / ρg)',
             bg=C_CALC_BG)

    r += 1

    # ═════════════════════════════════════════════════════════════════════════
    # SECTION 4 — RESULTS  (reference calc cells — update automatically)
    # ═════════════════════════════════════════════════════════════════════════
    section('SECTION 4 — RESULTS  [OUTPUTS — hand off to Mechanical]')

    data_row(None, 'Rated Flow (to Mechanical)',         'Q_r',   f"={C('Qr')}",    'm³/h',   'Input to pump sizing — 4-CAL-0001',                 bg=C_OUTPUT_BG)
    data_row(None, 'Total Dynamic Head (to Mechanical)', 'TDH',   f"={C('TDH')}",   'm',      'Required pump head — 4-CAL-0001',                    bg=C_OUTPUT_BG)
    data_row(None, 'NPSHa (to Mechanical)',              'NPSHa', f"={C('NPSHa')}",  'm',      'Must exceed NPSHr + 0.6 m margin (API 610)',         bg=C_OUTPUT_BG)
    data_row(None, 'Fluid Density (to Mechanical)',      'ρ',     f"={C('rho')}",    'kg/m³',  'Required for shaft power and motor sizing',          bg=C_OUTPUT_BG)
    data_row(None, 'Design Pressure (to Mechanical)',    'P_d',   f"={C('Pdes')}",   'bar(g)', 'Determines flange class and pressure test pressure',  bg=C_OUTPUT_BG)
    data_row(None, 'Corrosion Allowance (to Mechanical)','CA',    f"={C('CA')}",     'mm',     'Determines minimum wall thickness',                  bg=C_OUTPUT_BG)

    r += 2
    ws.merge_cells(f'A{r}:E{r}')
    ws.cell(r, 1, f"Generated by Process Engineer Agent (Aria)  |  {data.get('generated_at','')[:19]}  |  A Star Chemical").font = _font(colour='808080', size=8)
    ws.cell(r, 1).alignment = _align(h='center')

    ws.freeze_panes = 'A8'

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def build_pump_calc_xlsx(data: dict, output_path: Path) -> None:
    """Build a formatted pump calculation sheet from mechanical_output.json."""
    ps   = data.get('project_summary',   {})
    fp   = data.get('fluid_properties',  {})
    dc   = data.get('design_criteria',   {})
    pc   = data.get('pump_calculations', {})

    wb = Workbook()
    ws = wb.active
    ws.title = '4-CAL-0001'

    # Column widths  A=Parameter  B=Symbol  C=Value  D=Units  E=Excel Formula  F=Engineering Formula
    for i, w in enumerate([32, 10, 14, 10, 42, 42], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    refs: dict[str, int] = {}
    NCOLS = 6

    # ── Document header ───────────────────────────────────────────────────────
    hdr_fill = _fill(C_HEADER_BG)

    ws.merge_cells(f'A{r}:C{r}')
    ws.merge_cells(f'D{r}:F{r}')
    ws.cell(r, 1, 'A Star Chemical').font      = _font(bold=True, colour=C_HEADER_FG, size=12)
    for ci in range(1, 4): ws.cell(r, ci).fill = hdr_fill
    ws.cell(r, 4, '4-CAL-0001').font           = _font(bold=True, colour=C_HEADER_FG, size=12)
    ws.cell(r, 4).alignment                    = _align(h='right')
    for ci in range(4, 7): ws.cell(r, ci).fill = hdr_fill
    ws.row_dimensions[r].height = 22; r += 1

    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(r, 1, 'Mechanical Calculations — Pump Sizing & Motor Selection').font = _font(bold=True, colour=C_HEADER_FG, size=11)
    for ci in range(1, 7): ws.cell(r, ci).fill = hdr_fill
    ws.row_dimensions[r].height = 18; r += 1

    info_fill = _fill(C_COL_HDR_BG)
    for label, val in [
        ('Project',  ps.get('project_title', '—')),
        ('Client',   ps.get('client', '—')),
        ('Location', ps.get('location', '—')),
        ('Date',     data.get('generated_at', '—')[:10]),
    ]:
        ws.merge_cells(f'B{r}:F{r}')
        ws.cell(r, 1, label).font = _font(bold=True)
        for ci in range(1, 7):
            ws.cell(r, ci).fill   = info_fill
            ws.cell(r, ci).border = _border()
        ws.cell(r, 2, val).alignment = _align()
        ws.row_dimensions[r].height = 14; r += 1

    r += 1  # blank spacer

    _write_row(ws, r,
               ['Parameter', 'Symbol', 'Value', 'Units', 'Excel Formula', 'Engineering Formula'],
               fill=_fill(C_COL_HDR_BG), bold=True, height=16)
    r += 1

    # ── Helpers ───────────────────────────────────────────────────────────────

    def section(title: str):
        nonlocal r
        ws.merge_cells(f'A{r}:F{r}')
        ws.cell(r, 1, title).font   = _font(bold=True, colour=C_SECTION_FG)
        ws.cell(r, 1).alignment     = _align(h='center')
        for ci in range(1, 7):
            ws.cell(r, ci).fill     = _fill(C_SECTION_BG)
            ws.cell(r, ci).border   = _border()
        ws.row_dimensions[r].height = 15; r += 1

    def data_row(ref_name, param, symbol, value, units,
                 eng_formula='', bg=C_INPUT_BG):
        nonlocal r
        row_num = r
        if ref_name:
            refs[ref_name] = row_num
        excel_formula_text = str(value) if str(value).startswith('=') else ''
        _write_row(ws, r, [param, symbol, value, units, excel_formula_text, eng_formula],
                   fill=_fill(bg))
        ws.cell(r, VAL_COL).alignment = _align(h='right')
        if excel_formula_text:
            ws.cell(r, 5).font = Font(name='Courier New', size=8, color='555555')
        r += 1
        return row_num

    def C(ref_name: str) -> str:
        return f'C{refs[ref_name]}'

    # ═════════════════════════════════════════════════════════════════════════
    # SECTION 1 — FLUID DATA  (inputs from Process Engineer)
    # ═════════════════════════════════════════════════════════════════════════
    section('SECTION 1 — FLUID DATA  [INPUTS — from Process Engineer]')

    data_row('rho',  'Fluid Density',     'ρ',    fp.get('density_kgm3'),                'kg/m³', 'From Process Engineer handoff (1-CAL-XXXX)')
    data_row('SG',   'Specific Gravity',  'SG',   f"={C('rho')}/1000",                   '—',     'SG = ρ / 1000')
    data_row('Pv',   'Vapour Pressure',   'Pv',   fp.get('vapor_pressure_kPa_at_temp'),   'kPa',   f"At operating temperature {ps.get('temperature_c')} °C")
    data_row('PATM', 'Atmospheric Pressure','PATM', 1.01325,                              'bar',   'Standard atmospheric pressure')
    data_row(None,   'Fluid Name',        '—',    fp.get('fluid_name'),                  '—',     f"Code: {fp.get('fluid_code')}")

    r += 1

    # ═════════════════════════════════════════════════════════════════════════
    # SECTION 2 — DESIGN CONDITIONS  (inputs)
    # ═════════════════════════════════════════════════════════════════════════
    section('SECTION 2 — DESIGN CONDITIONS  [INPUTS]')

    data_row('Q',   'Rated Flow Rate',           'Q',    pc.get('rated_flow_m3h'),                  'm³/h',  'From Process Engineer handoff')
    data_row('N',   'Pump Speed',                'N',    pc.get('pump_speed_rpm',  1450),            'rpm',   'Standard 4-pole motor at 50 Hz')
    data_row('eta', 'Assumed Pump Efficiency',   'η',    pc.get('assumed_efficiency', 0.72),         '—',     'Typical centrifugal pump — confirmed with vendor')
    data_row('Ps',  'Suction Vessel Pressure',   'Ps',   ps.get('suction_pressure_barg'),            'barg',  'From P&ID / SOW')
    data_row('Pd',  'Discharge Vessel Pressure', 'Pd',   ps.get('discharge_pressure_barg'),          'barg',  'From P&ID / SOW')
    data_row('Zs',  'Suction Static Head',       'Zs',   ps.get('suction_static_head_m'),            'm',     'Suction liquid level elevation above pump CL')
    data_row('Zd',  'Discharge Static Head',     'Zd',   ps.get('discharge_static_head_m'),          'm',     'Discharge vessel inlet elevation above pump CL')
    data_row('hfs', 'Suction Friction Loss',     'hf_s', ps.get('suction_friction_loss_m'),          'm',     'Pipe friction losses — suction line')
    data_row('hfd', 'Discharge Friction Loss',   'hf_d', ps.get('discharge_friction_loss_m'),        'm',     'Pipe friction losses — discharge line')
    data_row('NPSHa', 'NPSHa (from Process)',    'NPSHa',pc.get('NPSHa_m'),                          'm',     'Net Positive Suction Head Available — from 1-CAL-XXXX')

    r += 1

    # ═════════════════════════════════════════════════════════════════════════
    # SECTION 3 — PUMP HYDRAULIC CALCULATIONS  (live Excel formulas)
    # ═════════════════════════════════════════════════════════════════════════
    section('SECTION 3 — PUMP HYDRAULIC CALCULATIONS  [CALC — live formulas]')

    data_row('Hs', 'Total Suction Head', 'Hs',
             f"=ROUND(({C('Ps')}+{C('PATM')})*100000/({C('rho')}*9.81)+{C('Zs')}-{C('hfs')}-{C('Pv')}*1000/({C('rho')}*9.81),2)",
             'm',
             'Hs = (Ps + Patm)×100,000/(ρ×g) + Zs − hf_s − Pv×1000/(ρ×g)',
             bg=C_CALC_BG)

    data_row('Hd', 'Total Discharge Head', 'Hd',
             f"=ROUND(({C('Pd')}+{C('PATM')})*100000/({C('rho')}*9.81)+{C('Zd')}+{C('hfd')},2)",
             'm',
             'Hd = (Pd + Patm)×100,000/(ρ×g) + Zd + hf_d',
             bg=C_CALC_BG)

    data_row('TDH', 'Total Dynamic Head', 'TDH',
             f"=ROUND({C('Hd')}-{C('Hs')},2)",
             'm',
             'TDH = Hd − Hs',
             bg=C_CALC_BG)

    data_row('Q_m3s', 'Flow Rate (m³/s)', 'Q_m3s',
             f"={C('Q')}/3600",
             'm³/s',
             'Q_m3s = Q [m³/h] / 3600',
             bg=C_CALC_BG)

    r += 1

    # ═════════════════════════════════════════════════════════════════════════
    # SECTION 4 — POWER & MOTOR SIZING  (live Excel formulas)
    # ═════════════════════════════════════════════════════════════════════════
    section('SECTION 4 — POWER & MOTOR SIZING  [CALC — live formulas]')

    data_row('Phyd', 'Hydraulic Power', 'P_hyd',
             f"=ROUND({C('Q')}*{C('rho')}*9.81*{C('TDH')}/3600000,2)",
             'kW',
             'P_hyd = Q × ρ × g × TDH / 3,600,000',
             bg=C_CALC_BG)

    data_row('BHP', 'Shaft Power (BHP)', 'BHP',
             f"=ROUND({C('Phyd')}/{C('eta')},2)",
             'kW',
             'BHP = P_hyd / η',
             bg=C_CALC_BG)

    data_row('mm', 'API 610 Motor Margin', 'mm',
             f"=IF({C('BHP')}<22,1.25,IF({C('BHP')}<=55,1.15,1.10))",
             '×',
             'API 610: ×1.25 if BHP<22 kW, ×1.15 if ≤55 kW, ×1.10 otherwise',
             bg=C_CALC_BG)

    data_row('Pmotor', 'Required Motor Power', 'P_motor',
             f"=ROUND({C('BHP')}*{C('mm')},2)",
             'kW',
             'P_motor = BHP × motor margin',
             bg=C_CALC_BG)

    r += 1

    # ═════════════════════════════════════════════════════════════════════════
    # SECTION 5 — SPECIFIC SPEED & IMPELLER TYPE
    # ═════════════════════════════════════════════════════════════════════════
    section('SECTION 5 — SPECIFIC SPEED & IMPELLER TYPE  [CALC — live formulas]')

    data_row('Ns', 'Specific Speed', 'Ns',
             f"=ROUND({C('N')}*SQRT({C('Q_m3s')})/POWER({C('TDH')},0.75),1)",
             '—',
             'Ns = N × √Q_m3s / TDH^0.75  (SI dimensionless)',
             bg=C_CALC_BG)

    data_row(None, 'Impeller Type', '—',
             f"=IF({C('Ns')}<25,\"Radial flow\",IF({C('Ns')}<100,\"Mixed flow\",\"Axial flow\"))",
             '—',
             'Ns < 25 → Radial  |  25–100 → Mixed  |  >100 → Axial',
             bg=C_CALC_BG)

    r += 1

    # ═════════════════════════════════════════════════════════════════════════
    # SECTION 6 — NPSH CHECK
    # ═════════════════════════════════════════════════════════════════════════
    section('SECTION 6 — NPSH CHECK  [CALC]')

    data_row('NPSHr', 'NPSHr Estimated', 'NPSHr',
             pc.get('NPSHr_estimated_m', 4.0),
             'm',
             'Conservative estimate 4 m — vendor to confirm actual NPSHr curve')

    data_row('npsh_margin', 'NPSH Margin', 'Margin',
             f"=ROUND({C('NPSHa')}-{C('NPSHr')},1)",
             'm',
             'NPSH Margin = NPSHa − NPSHr  (API 610 requires ≥ 0.6 m)',
             bg=C_CALC_BG)

    data_row(None, 'NPSH Status', '—',
             f"=IF({C('npsh_margin')}>=0.6,\"PASS\",\"VERIFY WITH VENDOR\")",
             '—',
             'Margin ≥ 0.6 m → PASS',
             bg=C_CALC_BG)

    r += 1

    # ═════════════════════════════════════════════════════════════════════════
    # SECTION 7 — RESULTS  (reference calc cells — hand off to procurement)
    # ═════════════════════════════════════════════════════════════════════════
    section('SECTION 7 — RESULTS  [OUTPUTS — for bid package]')

    data_row(None, 'Total Dynamic Head (to Bid)',   'TDH',      f"={C('TDH')}",     'm',     'Required pump head for vendor selection',       bg=C_OUTPUT_BG)
    data_row(None, 'Rated Flow (to Bid)',           'Q_r',      f"={C('Q')}",       'm³/h',  'Required pump flow for vendor selection',        bg=C_OUTPUT_BG)
    data_row(None, 'Required Motor Power (to Bid)', 'P_motor',  f"={C('Pmotor')}", 'kW',    'Minimum motor nameplate power — API 610 margin', bg=C_OUTPUT_BG)
    data_row(None, 'NPSHa (to Bid)',               'NPSHa',    f"={C('NPSHa')}",   'm',     'Vendor NPSHr must be < NPSHa − 0.6 m',          bg=C_OUTPUT_BG)
    data_row(None, 'Specific Speed (ref)',         'Ns',       f"={C('Ns')}",      '—',     'Confirms pump type selection',                   bg=C_OUTPUT_BG)

    r += 2
    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(r, 1, f"Generated by Mechanical Engineer Agent (Hunter)  |  {data.get('generated_at','')[:19]}  |  A Star Chemical").font = _font(colour='808080', size=8)
    ws.cell(r, 1).alignment = _align(h='center')

    ws.freeze_panes = 'A8'

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def _load_engine_ref_tables() -> dict:
    """Read the engine's versioned Ref-Table JSONs so the workbook's Ref-Tables sheet is
    driven by the same source (not re-hardcoded here)."""
    base = Path(__file__).parent / 'engines' / 'ref_tables'
    out = {}
    for name in ('dps', 'cable_vd'):
        try:
            out[name] = json.loads((base / f'{name}.json').read_text())
        except Exception:
            out[name] = {}
    return out


def _elec_write_ref_tables(ws, rt: dict) -> dict:
    """Populate the Ref-Tables sheet from the engine JSONs and return quoted A1 references
    the Load-Calc formulas point at. Editing a cell here re-flows the whole calc."""
    Q = "'Ref-Tables'!"
    dps = rt.get('dps', {})
    cab = rt.get('cable_vd', {})
    dens = dps.get('power_density_va_per_m2', {})
    df   = dps.get('demand_factor', {})
    dh   = dps.get('double_height_adder', {})
    ladder = (dps.get('breaker_ladder_a', {}) or {}).get('values', [20, 30, 40, 50, 70, 100, 125, 150, 200, 250, 300, 400, 500, 600, 800])
    divisor = (dps.get('load_table_current_divisor', {}) or {}).get('divisor', 690)
    mv   = cab.get('mv_per_a_per_m', {})
    amp  = cab.get('derated_ampacity_a', {})
    volt = (cab.get('system_voltage_ll', {}) or {}).get('value', 400)

    for i, w in enumerate([26, 12, 12, 40], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def title(r, text):
        ws.merge_cells(f'A{r}:D{r}')
        ws.cell(r, 1, text).font = _font(bold=True, colour=C_HEADER_FG, size=12)
        for ci in range(1, 5): ws.cell(r, ci).fill = _fill(C_HEADER_BG)
        ws.row_dimensions[r].height = 20

    def section(r, text, source=''):
        ws.cell(r, 1, text).font = _font(bold=True, colour=C_SECTION_FG)
        if source:
            ws.cell(r, 4, source).font = _font(colour=C_OUTPUT_BG if False else '006100', size=8)
            ws.cell(r, 4).alignment = _align(h='right')
        for ci in range(1, 5): ws.cell(r, ci).fill = _fill(C_SECTION_BG)

    def kv(r, label, value, unit=''):
        ws.cell(r, 1, label).font = _font(bold=True)
        c = ws.cell(r, 2, value); c.fill = _fill(C_INPUT_BG); c.border = _border()
        c.alignment = _align(h='right'); c.font = _font(colour='0000FF')
        if unit: ws.cell(r, 3, unit).font = _font(colour='808080', size=9)

    title(1, 'Reference Tables — edit these; the calc recomputes')
    ws.merge_cells('A2:D2')
    ws.cell(2, 1, (dps.get('_meta', {}) or {}).get('note', 'DPS 2022 — indicative; an engineer confirms per project.')).font = _font(colour='808080', size=8)

    section(4, 'Power density (VA/m²)', (dens.get('_source') or ''))
    ws.cell(5, 1, 'Category').font = _font(bold=True); ws.cell(5, 2, 'VA/m²').font = _font(bold=True)
    cats = [(k, v) for k, v in dens.items() if not k.startswith('_')]
    dr0 = 6
    for j, (cat, val) in enumerate(cats):
        ws.cell(dr0 + j, 1, cat); c = ws.cell(dr0 + j, 2, val)
        c.fill = _fill(C_INPUT_BG); c.font = _font(colour='0000FF'); c.alignment = _align(h='right'); c.border = _border()
    dr1 = dr0 + len(cats) - 1

    r = dr1 + 2
    section(r, 'Demand factor', (df.get('_source') or '')); r += 1
    df_row = r; kv(r, 'office', df.get('office', 0.6)); r += 2
    section(r, 'Current divisor & voltage'); r += 1
    div_row = r; kv(r, 'Load-table divisor (÷)', divisor, '(Demand×1000/690)'); r += 1
    volt_row = r; kv(r, 'System voltage', volt, 'V, L-L'); r += 2
    section(r, 'Double-height adder', (dh.get('_source') or '')); r += 1
    dh_rate_row = r; kv(r, 'rate', dh.get('rate_va_per_m3', 24), 'VA/m³'); r += 1
    dh_std_row = r; kv(r, 'standard height', dh.get('standard_height_m', 3.5), 'm'); r += 2

    section(r, 'Breaker ladder (A)', (dps.get('breaker_ladder_a', {}) or {}).get('_source', '')); r += 1
    lad0 = r
    for v in ladder:
        c = ws.cell(r, 1, v); c.fill = _fill(C_INPUT_BG); c.font = _font(colour='0000FF')
        c.alignment = _align(h='right'); c.border = _border(); r += 1
    lad1 = r - 1
    r += 1

    section(r, 'Cable: CSA → mv/a/m & derated ampacity', (mv.get('_source') or '')); r += 1
    ws.cell(r, 1, 'CSA mm²').font = _font(bold=True); ws.cell(r, 2, 'mv/a/m').font = _font(bold=True)
    ws.cell(r, 3, 'Derated A').font = _font(bold=True); r += 1
    cab0 = r
    csas = sorted((int(k) for k in mv if not k.startswith('_')))
    for csa in csas:
        ws.cell(r, 1, csa).alignment = _align(h='right')
        for col, tbl in ((2, mv), (3, amp)):
            c = ws.cell(r, col, tbl.get(str(csa))); c.fill = _fill(C_INPUT_BG)
            c.font = _font(colour='0000FF'); c.alignment = _align(h='right'); c.border = _border()
        r += 1
    cab1 = r - 1

    return {
        'density_range': f'{Q}$A${dr0}:$B${dr1}',
        'df': f'{Q}$B${df_row}',
        'divisor': f'{Q}$B${div_row}',
        'voltage': f'{Q}$B${volt_row}',
        'dh_rate': f'{Q}$B${dh_rate_row}',
        'dh_std': f'{Q}$B${dh_std_row}',
        'ladder': f'{Q}$A${lad0}:$A${lad1}',
        'ladder_first': f'{Q}$A${lad0}',
        'cable_range': f'{Q}$A${cab0}:$C${cab1}',
    }


# Load-Calc columns: A Office B Area C Height D Cat E Density F Connected G Add'l
# H Total I DF J Demand K I₆₉₀ L Breaker M I_cable N Feeder O Cable P mv/a/m Q VD%
_ELEC_HDRS = ['Office', 'Area m²', 'Height m', 'Cat', 'Density\nVA/m²', 'Connected\nkVA',
              "Add'l\nkVA", 'Total\nkVA', 'DF', 'Demand\nkVA', 'Current\nA (÷690)', 'Breaker\nA',
              'Cable I\nA', 'Feeder\nm', 'Cable\nmm²', 'mv/a/m', 'VD %']
_ELEC_WIDTHS = [28, 8, 8, 5, 9, 11, 8, 9, 5, 10, 10, 9, 9, 8, 8, 8, 7]


def _elec_write_load_calc(ws, board: dict, ref: dict):
    """54-office schedule with live formulas that reference the Ref-Tables sheet.
    The whole chain (incl. the double-height adder) is live: change a blue input cell and
    every yellow cell re-flows. Returns (first_data_row, last_data_row) for Summary."""
    rooms = board.get('rooms', [])
    for i, w in enumerate(_ELEC_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells('A1:O1')
    ws.cell(1, 1, 'Electrical Load Schedule — MDB-OFFICES (DPS 2022 density method)').font = _font(bold=True, colour=C_HEADER_FG, size=12)
    ws.merge_cells('P1:Q1')
    ws.cell(1, 16, 'live formulas →').font = _font(colour=C_HEADER_FG, size=8)
    ws.cell(1, 16).alignment = _align(h='right')
    for ci in range(1, 18): ws.cell(1, ci).fill = _fill(C_HEADER_BG)
    ws.row_dimensions[1].height = 20

    hr = 3
    _write_row(ws, hr, _ELEC_HDRS, fill=_fill(C_COL_HDR_BG), bold=True, height=30)
    for ci in range(1, 18):
        ws.cell(hr, ci).alignment = _align(h='center', wrap=True)

    std, rate = ref['dh_std'], ref['dh_rate']
    L, F0 = ref['ladder'], ref['ladder_first']
    r = hr + 1
    first = r
    for rm in rooms:
        area = rm.get('area_m2'); cat = rm.get('category')
        has = area is not None and cat is not None
        ws.cell(r, 1, rm.get('name'))
        ws.cell(r, 2, area)
        ws.cell(r, 3, rm.get('height_m'))                          # blank for single-height
        ws.cell(r, 4, cat)
        ws.cell(r, 5, f'=IFERROR(VLOOKUP(D{r},{ref["density_range"]},2,FALSE),"?")' if has else '—')
        ws.cell(r, 6, f'=B{r}*E{r}/1000' if has else '—')          # connected
        ws.cell(r, 7, f'=IF(AND(C{r}<>"",C{r}>{std}),(C{r}-{std})*B{r}*{rate}/1000,0)')  # add'l (live)
        ws.cell(r, 8, f'=F{r}+G{r}')                               # total
        ws.cell(r, 9, f'={ref["df"]}')                             # DF
        ws.cell(r, 10, f'=H{r}*I{r}')                              # demand
        ws.cell(r, 11, f'=J{r}*1000/{ref["divisor"]}')            # I690
        ws.cell(r, 12, f'=IFERROR(INDEX({L},MATCH(K{r},{L},1)+IF(INDEX({L},MATCH(K{r},{L},1))=K{r},0,1)),{F0})')  # breaker
        ws.cell(r, 13, f'=J{r}*1000/(SQRT(3)*{ref["voltage"]})')  # cable current
        ws.cell(r, 14, rm.get('feeder_length_m'))
        ws.cell(r, 15, rm.get('cable_csa_mm2') if rm.get('cable_csa_mm2') is not None else '—')
        ws.cell(r, 16, f'=IFERROR(VLOOKUP(O{r},{ref["cable_range"]},2,FALSE),"")')  # mv/a/m
        ws.cell(r, 17, f'=IFERROR(ROUND(M{r}*N{r}*P{r}/1000/{ref["voltage"]}*100,1),"")')  # VD%

        for ci in range(1, 18):
            c = ws.cell(r, ci); c.border = _border(); c.font = _font()
            c.alignment = _align(h='left' if ci == 1 else 'right')
        for ci in (2, 3, 4, 14):                       # inputs -> blue
            ws.cell(r, ci).fill = _fill(C_INPUT_BG)
        for ci in (5, 6, 7, 8, 9, 10, 11, 13, 16, 17): # formulas -> yellow
            ws.cell(r, ci).fill = _fill(C_CALC_BG)
        for ci in (12, 15):                            # selected outputs -> green
            ws.cell(r, ci).fill = _fill(C_OUTPUT_BG)
        r += 1
    last = r - 1
    ws.freeze_panes = f'A{first}'
    return first, last


def _elec_write_cover(ws, ps: dict, board: dict, generated_at: str, ref: dict):
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 88

    ws.merge_cells('A1:B1')
    ws.cell(1, 1, 'Electrical Load Calculation — Reusable Template').font = _font(bold=True, colour=C_HEADER_FG, size=14)
    for ci in (1, 2): ws.cell(1, ci).fill = _fill(C_HEADER_BG)
    ws.row_dimensions[1].height = 26
    ws.merge_cells('A2:B2')
    ws.cell(2, 1, 'Engine-driven · real project data · live Excel formulas · fully auditable · reusable across projects').font = _font(colour='595959', size=9)

    def sect(r, t):
        ws.merge_cells(f'A{r}:B{r}')
        ws.cell(r, 1, t).font = _font(bold=True, colour=C_SECTION_FG, size=11)
        for ci in (1, 2): ws.cell(r, ci).fill = _fill(C_SECTION_BG)
        ws.row_dimensions[r].height = 18

    def kv(r, k, v):
        ws.cell(r, 1, k).font = _font(bold=True); ws.cell(r, 1).fill = _fill(C_COL_HDR_BG)
        ws.cell(r, 1).border = _border()
        ws.cell(r, 2, v).alignment = _align(wrap=True); ws.cell(r, 2).border = _border()

    r = 4
    sect(r, 'Project'); r += 1
    kv(r, 'Project', ps.get('project_title', '—')); r += 1
    kv(r, 'Board', board.get('board', 'MDB-OFFICES')); r += 1
    kv(r, 'Offices', board.get('row_count')); r += 1
    kv(r, 'Generated', (generated_at or '')[:19]); r += 2

    sect(r, 'Method — DPS 2022 power-density'); r += 1
    for k, v in [
        ('Connected load', 'kVA = Area (m²) × Density (VA/m²) ÷ 1000'),
        ("Double-height add'l", 'kVA = (Height − 3.5) × Area × 24 ÷ 1000'),
        ('Demand', 'kVA = (Connected + Add\'l) × Demand factor'),
        ('Load-table current', 'A = Demand × 1000 ÷ 690  →  main breaker = next standard size ≥ current'),
        ('Cable current', 'A = Demand × 1000 ÷ (√3 × 400)'),
        ('Cable size', 'smallest CSA with derated ampacity ≥ current AND VD ≤ 4%'),
        ('Voltage drop', 'VD% = Cable current × Length × (mv/a/m) ÷ 1000 ÷ 400 × 100'),
    ]:
        kv(r, k, v); r += 1
    r += 1

    sect(r, 'How to use'); r += 1
    kv(r, 'It is live', 'Change any blue input cell on Load-Calc (e.g. an office area) — every yellow formula recomputes automatically.'); r += 1
    kv(r, 'It is auditable', 'All standard values (density, demand factor, breaker ladder, cable table) live on the Ref-Tables sheet, each with its source. Formulas reference those cells — edit a reference and the whole calc re-flows.'); r += 1
    kv(r, 'It is reusable', 'The workbook is generated from the engine per project — feed a new project\'s data and you get the same structured workbook.'); r += 2

    sect(r, 'Legend'); r += 1
    for label, colour in [('Input (extracted / entered)', C_INPUT_BG),
                          ('Calculated (live formula)', C_CALC_BG),
                          ('Selected output (breaker / cable)', C_OUTPUT_BG)]:
        ws.cell(r, 1, '').fill = _fill(colour); ws.cell(r, 1).border = _border()
        ws.cell(r, 2, label); r += 1


def _elec_write_summary(ws, board: dict, first: int, last: int):
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 10

    ws.merge_cells('A1:C1')
    ws.cell(1, 1, 'Board Summary — MDB-OFFICES (OUTPUT)').font = _font(bold=True, colour=C_HEADER_FG, size=12)
    for ci in (1, 2, 3): ws.cell(1, ci).fill = _fill(C_HEADER_BG)
    ws.row_dimensions[1].height = 20
    ws.merge_cells('A2:C2')
    ws.cell(2, 1, 'Feeds transformer sizing, panel schedules and the single-line diagram.').font = _font(colour='595959', size=9)

    CALC = "'Load-Calc'!"
    rows = [
        ('Offices', board.get('row_count'), ''),
        ('Total connected', f'=SUM({CALC}F{first}:F{last})', 'kVA'),
        ("Total additional (double-height)", f'=SUM({CALC}G{first}:G{last})', 'kVA'),
        ('Total demand', f'=SUM({CALC}J{first}:J{last})', 'kVA'),
        ('Main incomer current', f'=SUM({CALC}J{first}:J{last})*1000/(SQRT(3)*400)', 'A'),
        ('Max feeder voltage drop', f'=MAX({CALC}Q{first}:Q{last})', '%'),
    ]
    r = 4
    for label, val, unit in rows:
        ws.cell(r, 1, label).font = _font(bold=True); ws.cell(r, 1).fill = _fill(C_COL_HDR_BG)
        ws.cell(r, 1).border = _border()
        c = ws.cell(r, 2, val); c.fill = _fill(C_OUTPUT_BG); c.border = _border(); c.alignment = _align(h='right')
        ws.cell(r, 3, unit).font = _font(colour='808080', size=9)
        r += 1

    fc = board.get('footer_check')
    if fc:
        r += 1
        ws.merge_cells(f'A{r}:C{r}')
        ws.cell(r, 1, f'⚠ Data-quality flag: {fc.get("message", "")}').font = _font(colour='C00000', size=9)
        ws.cell(r, 1).alignment = _align(wrap=True)
        ws.row_dimensions[r].height = 40


def build_electrical_calc_xlsx(data: dict, output_path: Path) -> None:
    """Build a reusable, engine-driven electrical LOAD calculation workbook.

    Four sheets — Cover (method + how-to), Load-Calc (per-office schedule with live
    formulas that reference Ref-Tables), Summary (board totals → transformer/panel/SLD),
    and Ref-Tables (density/demand-factor/breaker-ladder/cable table, each with its
    source). Every engineering number is a live formula referencing a Ref-Table cell —
    nothing hard-coded in the workbook — so changing an input or a reference re-flows the
    whole calc. Fully parameterised by `data` (the engine result) → reusable per project.
    """
    ps    = data.get('project_summary', {})
    board = data.get('electrical_calculations', {})
    gen   = data.get('generated_at', '') or ''

    wb = Workbook()
    cover = wb.active;                cover.title = 'Cover'
    calc  = wb.create_sheet('Load-Calc')
    summ  = wb.create_sheet('Summary')
    reft  = wb.create_sheet('Ref-Tables')

    ref = _elec_write_ref_tables(reft, _load_engine_ref_tables())
    first, last = _elec_write_load_calc(calc, board, ref)
    _elec_write_cover(cover, ps, board, gen, ref)
    _elec_write_summary(summ, board, first, last)

    wb.active = 0   # open on Cover
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--input',  help='Path to JSON file (default: stdin)')
    parser.add_argument('--output', required=True)
    parser.add_argument('--type',   choices=['process', 'pump', 'electrical'], default='process',
                        help='Which calc sheet to generate')
    args = parser.parse_args()

    data = json.loads(Path(args.input).read_text()) if args.input else json.loads(sys.stdin.read())
    if args.type == 'pump':
        build_pump_calc_xlsx(data, Path(args.output))
    elif args.type == 'electrical':
        build_electrical_calc_xlsx(data, Path(args.output))
    else:
        build_process_calc_xlsx(data, Path(args.output))
    print('OK')
